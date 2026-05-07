"""
HR 财务社保人员统计表
工作簿结构:
  Sheet1 花名册     — 员工基本信息
  Sheet2 薪酬统计   — 工资明细 + 个税
  Sheet3 社保明细   — 五险一金明细
  Sheet4 部门汇总   — SUMIF 汇总 + 格式
"""
import sys, os, json, time, tempfile
sys.path.insert(0, os.path.dirname(__file__))
os.environ.setdefault("EXCEL_SNAPSHOT_DIR", os.path.join(os.path.dirname(__file__), ".snapshots"))
os.environ.setdefault("EXCEL_MAX_SNAPSHOTS", "50")

import openpyxl

from excel_mcp.tools.file_tools   import open_workbook, save_workbook, close_workbook
from excel_mcp.tools.range_tools  import (
    set_values, set_formulas, format_range, set_number_format,
    merge_cells, autofit, set_row_height, set_column_width,
    freeze_panes,
)
from excel_mcp.tools.sheet_tools  import create_sheet, rename_sheet, set_tab_color
from excel_mcp.tools.named_range_tools import create_named_range

results = []
def ck(label, raw):
    d = json.loads(raw) if isinstance(raw, str) else raw
    ok = d.get("success", False)
    print(f"  {'OK' if ok else 'NG'}  {label}")
    if not ok:
        print(f"       ERR: {d.get('error', d)}")
    results.append((label, ok))
    return d

# ── Kill stale WPS & create workbook ──────────────────────────────────────────
import subprocess as _sp
_sp.run(["taskkill","/f","/im","et.exe"],  capture_output=True)
_sp.run(["taskkill","/f","/im","wps.exe"], capture_output=True)
time.sleep(4)

OUT = os.path.join(tempfile.gettempdir(), "HR社保财务统计.xlsx")
wb0 = openpyxl.Workbook(); wb0.active.title = "Sheet1"; wb0.save(OUT)
print(f"\n[ 准备 ]  {OUT}")
ck("open_workbook", open_workbook(OUT))
time.sleep(1)

ck("rename Sheet1 -> 花名册", rename_sheet("Sheet1", "花名册"))
ck("create 薪酬统计",         create_sheet("薪酬统计"))
ck("create 社保明细",         create_sheet("社保明细"))
ck("create 部门汇总",         create_sheet("部门汇总"))

# ══════════════════════════════════════════════════════════════════════════════
# 1. 花名册
# ══════════════════════════════════════════════════════════════════════════════
print("\n[ 1 ] 花名册")

SH = "花名册"
ck("tab color", set_tab_color(SH, "#4472C4"))

ck("title merge",  merge_cells("A1:H1", sheet=SH))
ck("title text",   set_values("A1", [["XX 科技有限公司 — 员工花名册（2025年）"]], sheet=SH))
ck("title format", format_range("A1", sheet=SH,
    bold=True, font_size=14, h_align="center", v_align="center",
    bg_color="#1F3864", font_color="#FFFFFF"))
ck("row1 height",  set_row_height(1, 36, sheet=SH))

HEADERS_ROSTER = [["编号","姓名","部门","岗位","性别","入职日期","合同类型","状态"]]
ck("headers", set_values("A2", HEADERS_ROSTER, sheet=SH))
ck("header format", format_range("A2:H2", sheet=SH,
    bold=True, bg_color="#2E75B6", font_color="#FFFFFF",
    h_align="center", border="outer"))

ROSTER = [
    ["E001","张明","研发部","技术总监","男","2019-03-01","无固定期限","在职"],
    ["E002","李娜","研发部","高级工程师","女","2020-06-15","固定期限","在职"],
    ["E003","王强","研发部","工程师","男","2021-09-01","固定期限","在职"],
    ["E004","赵磊","研发部","工程师","男","2022-03-15","固定期限","在职"],
    ["E005","陈静","市场部","市场总监","女","2018-07-01","无固定期限","在职"],
    ["E006","刘洋","市场部","市场经理","男","2020-01-10","固定期限","在职"],
    ["E007","孙芳","市场部","市场专员","女","2022-08-01","固定期限","在职"],
    ["E008","周涛","财务部","财务总监","男","2017-05-01","无固定期限","在职"],
    ["E009","吴梅","财务部","会计","女","2021-04-01","固定期限","在职"],
    ["E010","郑华","财务部","出纳","女","2023-01-10","固定期限","在职"],
    ["E011","冯丽","人事部","人事总监","女","2019-09-01","无固定期限","在职"],
    ["E012","陈刚","人事部","招聘专员","男","2022-06-01","固定期限","在职"],
    ["E013","褚燕","运营部","运营总监","女","2020-02-01","无固定期限","在职"],
    ["E014","卫东","运营部","运营专员","男","2022-11-01","固定期限","在职"],
    ["E015","蒋涛","运营部","运营专员","男","2023-05-15","固定期限","在职"],
]
ck("data", set_values("A3", ROSTER, sheet=SH))

# Alternating row colors
for i, row_idx in enumerate(range(3, 3+len(ROSTER))):
    color = "#DEEAF1" if i % 2 == 0 else "#FFFFFF"
    ck(f"row{row_idx} color", format_range(
        f"A{row_idx}:H{row_idx}", sheet=SH, bg_color=color, border="outer"))

ck("border all",   format_range("A2:H17", sheet=SH, border="all"))
ck("center cols",  format_range("A2:H17", sheet=SH, h_align="center"))
ck("autofit",      autofit("A:H", sheet=SH))
ck("freeze row1",  freeze_panes(2, 0, sheet=SH))

# ══════════════════════════════════════════════════════════════════════════════
# 2. 薪酬统计
# ══════════════════════════════════════════════════════════════════════════════
print("\n[ 2 ] 薪酬统计")
SH = "薪酬统计"
ck("tab color", set_tab_color(SH, "#ED7D31"))

ck("title merge",  merge_cells("A1:K1", sheet=SH))
ck("title text",   set_values("A1", [["XX 科技有限公司 — 2025年薪酬统计表"]], sheet=SH))
ck("title format", format_range("A1", sheet=SH,
    bold=True, font_size=14, h_align="center", v_align="center",
    bg_color="#843C0C", font_color="#FFFFFF"))
ck("row1 height",  set_row_height(1, 36, sheet=SH))

SALARY_HEADERS = [[
    "编号","姓名","部门","基本工资","岗位工资","绩效工资","餐补/交补","应发合计",
    "社保个人","个人所得税","实发工资"
]]
ck("headers", set_values("A2", SALARY_HEADERS, sheet=SH))
ck("header format", format_range("A2:K2", sheet=SH,
    bold=True, bg_color="#C55A11", font_color="#FFFFFF", h_align="center"))

# Base salary data (cols A-G; cols H-K will be formulas)
# 社保个人负担: 养老8%+医疗2%+失业0.5% ≈ 10.5% of base，在社保明细表会精确计算，这里近似
SALARY_DATA = [
    ["E001","张明","研发部", 15000, 8000, 6000, 500],
    ["E002","李娜","研发部", 12000, 5000, 4500, 500],
    ["E003","王强","研发部", 10000, 3000, 3000, 500],
    ["E004","赵磊","研发部",  9000, 2500, 2500, 500],
    ["E005","陈静","市场部", 14000, 7000, 8000, 500],
    ["E006","刘洋","市场部", 10000, 4000, 5000, 500],
    ["E007","孙芳","市场部",  7000, 2000, 2000, 500],
    ["E008","周涛","财务部", 13000, 6000, 5000, 500],
    ["E009","吴梅","财务部",  8000, 2000, 2000, 500],
    ["E010","郑华","财务部",  6500, 1500, 1500, 500],
    ["E011","冯丽","人事部", 12000, 5000, 4000, 500],
    ["E012","陈刚","人事部",  7500, 2000, 2000, 500],
    ["E013","褚燕","运营部", 11000, 5000, 5000, 500],
    ["E014","卫东","运营部",  7000, 2000, 2000, 500],
    ["E015","蒋涛","运营部",  7000, 2000, 1500, 500],
]
ck("salary data", set_values("A3", SALARY_DATA, sheet=SH))

# Formulas: H=应发合计, I=社保个人(近似10.5%), J=个税(超额累进简化), K=实发
# 个税 = MAX(0, (H - I - 5000) * 税率 - 速算扣除数)  用简化公式
salary_formulas = []
for i in range(len(SALARY_DATA)):
    r = i + 3
    h = f"=D{r}+E{r}+F{r}+G{r}"
    i_col = f"=ROUND(D{r}*0.105,2)"         # 社保个人约10.5%基本工资
    # 简化个税：应纳税所得额 = 应发 - 社保个人 - 5000起征点
    # 用 MAX(0,...) 保证不为负
    tax = (
        f"=MAX(0,ROUND(IF(H{r}-{i_col[1:]}-5000<=0,0,"
        f"IF(H{r}-{i_col[1:]}-5000<=36000,(H{r}-{i_col[1:]}-5000)*0.03,"
        f"IF(H{r}-{i_col[1:]}-5000<=144000,(H{r}-{i_col[1:]}-5000)*0.1-2520,"
        f"(H{r}-{i_col[1:]}-5000)*0.2-16920))),2))"
    )
    k = f"=H{r}-I{r}-J{r}"
    salary_formulas.append([h, i_col, tax, k])

ck("应发合计 formulas",   set_formulas("H3", [[r[0]] for r in salary_formulas], sheet=SH))
ck("社保个人 formulas",   set_formulas("I3", [[r[1]] for r in salary_formulas], sheet=SH))
ck("个税 formulas",       set_formulas("J3", [[r[2]] for r in salary_formulas], sheet=SH))
ck("实发工资 formulas",   set_formulas("K3", [[r[3]] for r in salary_formulas], sheet=SH))

# 汇总行
last = 2 + len(SALARY_DATA)
ck("total label",  set_values(f"C{last+1}", [["合  计"]], sheet=SH))
for col_letter in ["D","E","F","G","H","I","J","K"]:
    ck(f"total {col_letter}",
       set_formulas(f"{col_letter}{last+1}",
                    [[f"=SUM({col_letter}3:{col_letter}{last})"]], sheet=SH))
ck("total bold",   format_range(f"A{last+1}:K{last+1}", sheet=SH,
                                bold=True, bg_color="#FCE4D6"))

# Number format
for col_letter in ["D","E","F","G","H","I","J","K"]:
    ck(f"acct fmt {col_letter}", set_number_format(
        f"{col_letter}3:{col_letter}{last+1}", "#,##0.00", sheet=SH))

ck("border",    format_range(f"A2:K{last+1}", sheet=SH, border="all"))
ck("center AC", format_range(f"A2:C{last+1}", sheet=SH, h_align="center"))
ck("autofit",   autofit("A:K", sheet=SH))
ck("freeze",    freeze_panes(2, 3, sheet=SH))

# ══════════════════════════════════════════════════════════════════════════════
# 3. 社保明细
# ══════════════════════════════════════════════════════════════════════════════
print("\n[ 3 ] 社保明细")
SH = "社保明细"
ck("tab color", set_tab_color(SH, "#70AD47"))

ck("title merge",  merge_cells("A1:P1", sheet=SH))
ck("title text",   set_values("A1", [["XX 科技有限公司 — 2025年社保公积金明细"]], sheet=SH))
ck("title format", format_range("A1", sheet=SH,
    bold=True, font_size=14, h_align="center", v_align="center",
    bg_color="#375623", font_color="#FFFFFF"))
ck("row1 height",  set_row_height(1, 36, sheet=SH))

# 第二行：大组标题
ck("group headers merge 养老", merge_cells("D2:E2", sheet=SH))
ck("group headers merge 医疗", merge_cells("F2:G2", sheet=SH))
ck("group headers merge 失业", merge_cells("H2:I2", sheet=SH))
ck("group headers merge 公积金", merge_cells("M2:N2", sheet=SH))
GROUP_HEADERS = [["编号","姓名","部门","养老保险","","医疗保险","","失业保险","",
                  "工伤保险","生育保险","","住房公积金","","社保合计","公积金合计"]]
ck("group row", set_values("A2", GROUP_HEADERS, sheet=SH))

# 第三行：明细标题
DETAIL_HEADERS = [[
    "编号","姓名","部门",
    "养老(单位16%)","养老(个人8%)",
    "医疗(单位8%)","医疗(个人2%)",
    "失业(单位0.5%)","失业(个人0.5%)",
    "工伤(单位0.4%)","生育(单位0.8%)","",
    "公积金(单位12%)","公积金(个人12%)",
    "社保合计(个人)","公积金合计(个人)"
]]
ck("detail headers", set_values("A3", DETAIL_HEADERS, sheet=SH))

ck("header format row2", format_range("A2:P2", sheet=SH,
    bold=True, bg_color="#375623", font_color="#FFFFFF", h_align="center"))
ck("header format row3", format_range("A3:P3", sheet=SH,
    bold=True, bg_color="#548235", font_color="#FFFFFF", h_align="center"))
ck("row2 height", set_row_height(2, 22, sheet=SH))
ck("row3 height", set_row_height(3, 30, sheet=SH))

# 缴费基数 = 基本工资（从薪酬统计表 D 列引用）
# 各险种金额公式（引用薪酬统计!D列）
SS_BASE_SALARIES = [15000,12000,10000,9000,14000,10000,7000,13000,8000,6500,12000,7500,11000,7000,7000]
NAMES = [r[1] for r in SALARY_DATA]
DEPTS = [r[2] for r in SALARY_DATA]
IDS   = [r[0] for r in SALARY_DATA]

ss_data = []
for base in SS_BASE_SALARIES:
    # 缴费基数上限/下限处理（简化：直接用基本工资）
    e_pension   = round(base * 0.16, 2)   # 养老单位
    p_pension   = round(base * 0.08, 2)   # 养老个人
    e_medical   = round(base * 0.08, 2)   # 医疗单位
    p_medical   = round(base * 0.02, 2)   # 医疗个人
    e_unemp     = round(base * 0.005, 2)  # 失业单位
    p_unemp     = round(base * 0.005, 2)  # 失业个人
    e_injury    = round(base * 0.004, 2)  # 工伤单位
    e_maternity = round(base * 0.008, 2)  # 生育单位
    e_hfund     = round(base * 0.12, 2)   # 公积金单位
    p_hfund     = round(base * 0.12, 2)   # 公积金个人
    ss_data.append([
        e_pension, p_pension,
        e_medical, p_medical,
        e_unemp,   p_unemp,
        e_injury,  e_maternity, "",
        e_hfund,   p_hfund,
    ])

# Write ID/Name/Dept
ck("id name dept", set_values("A4", [[IDS[i], NAMES[i], DEPTS[i]] for i in range(15)], sheet=SH))
# Write insurance amounts
ck("ss amounts", set_values("D4", ss_data, sheet=SH))

# 社保合计(个人) = 养老个人 + 医疗个人 + 失业个人  (cols E + G + I)
# 公积金合计(个人) = 公积金个人 (col N)
ss_formulas_o = []
ss_formulas_p = []
for i in range(15):
    r = i + 4
    ss_formulas_o.append([f"=E{r}+G{r}+I{r}"])   # 社保合计(个人)
    ss_formulas_p.append([f"=N{r}"])               # 公积金合计(个人)
ck("社保个人合计 formula", set_formulas("O4", ss_formulas_o, sheet=SH))
ck("公积金个人合计 formula", set_formulas("P4", ss_formulas_p, sheet=SH))

# 合计行
last_ss = 3 + 15
ck("ss total label", set_values(f"C{last_ss+1}", [["合  计"]], sheet=SH))
for col_l in list("DEFGHIJKMNO") + ["N","O","P"]:
    try:
        ck(f"ss total {col_l}", set_formulas(f"{col_l}{last_ss+1}",
            [[f"=SUM({col_l}4:{col_l}{last_ss})"]], sheet=SH))
    except Exception:
        pass
ck("ss total bold", format_range(f"A{last_ss+1}:P{last_ss+1}", sheet=SH,
                                 bold=True, bg_color="#E2EFDA"))

ck("ss number format", set_number_format(f"D4:P{last_ss+1}", "#,##0.00", sheet=SH))
ck("ss border",     format_range(f"A2:P{last_ss+1}", sheet=SH, border="all"))
ck("ss center ABC", format_range(f"A2:C{last_ss+1}", sheet=SH, h_align="center"))
ck("ss autofit",    autofit("A:P", sheet=SH))
ck("ss freeze",     freeze_panes(3, 3, sheet=SH))

# ══════════════════════════════════════════════════════════════════════════════
# 4. 部门汇总
# ══════════════════════════════════════════════════════════════════════════════
print("\n[ 4 ] 部门汇总")
SH = "部门汇总"
ck("tab color", set_tab_color(SH, "#7030A0"))

ck("title merge",  merge_cells("A1:I1", sheet=SH))
ck("title text",   set_values("A1", [["XX 科技有限公司 — 2025年部门人员薪酬汇总"]], sheet=SH))
ck("title format", format_range("A1", sheet=SH,
    bold=True, font_size=14, h_align="center", v_align="center",
    bg_color="#4B0082", font_color="#FFFFFF"))
ck("row1 height",  set_row_height(1, 36, sheet=SH))

DEPT_HEADERS = [["部门","人数","应发合计","社保个人合计","个税合计","实发合计","社保公司合计","公积金个人","公积金公司"]]
ck("dept headers", set_values("A2", DEPT_HEADERS, sheet=SH))
ck("header format", format_range("A2:I2", sheet=SH,
    bold=True, bg_color="#7030A0", font_color="#FFFFFF", h_align="center"))

DEPTS_UNIQ = ["研发部","市场部","财务部","人事部","运营部"]
dept_data = [[d] for d in DEPTS_UNIQ]
ck("dept names", set_values("A3", dept_data, sheet=SH))

# SUMIF formulas referencing 薪酬统计 and 社保明细
dept_formulas = []
for i, dept in enumerate(DEPTS_UNIQ):
    r = i + 3
    dept_ref = dept  # Used in formula string
    # 薪酬统计: C=部门, H=应发, I=社保个人, J=个税, K=实发
    # 社保明细: C=部门, D=养老单位, F=医疗单位, H=失业单位, J=工伤, K=生育, M=公积金单位, N=公积金个人
    cnt   = f'=COUNTIF(薪酬统计!C:C,A{r})'
    yfhj  = f'=SUMIF(薪酬统计!C:C,A{r},薪酬统计!H:H)'
    sbyg  = f'=SUMIF(薪酬统计!C:C,A{r},薪酬统计!I:I)'
    gs    = f'=SUMIF(薪酬统计!C:C,A{r},薪酬统计!J:J)'
    sfgz  = f'=SUMIF(薪酬统计!C:C,A{r},薪酬统计!K:K)'
    # 公司社保: D+F+H+J+K from 社保明细
    sbgs  = (f'=SUMIF(社保明细!C:C,A{r},社保明细!D:D)'
             f'+SUMIF(社保明细!C:C,A{r},社保明细!F:F)'
             f'+SUMIF(社保明细!C:C,A{r},社保明细!H:H)'
             f'+SUMIF(社保明细!C:C,A{r},社保明细!J:J)'
             f'+SUMIF(社保明细!C:C,A{r},社保明细!K:K)')
    gjjyg = f'=SUMIF(社保明细!C:C,A{r},社保明细!N:N)'
    gjjgs = f'=SUMIF(社保明细!C:C,A{r},社保明细!M:M)'
    dept_formulas.append([cnt, yfhj, sbyg, gs, sfgz, sbgs, gjjyg, gjjgs])

ck("人数",     set_formulas("B3", [[r[0]] for r in dept_formulas], sheet=SH))
ck("应发合计", set_formulas("C3", [[r[1]] for r in dept_formulas], sheet=SH))
ck("社保个人", set_formulas("D3", [[r[2]] for r in dept_formulas], sheet=SH))
ck("个税",     set_formulas("E3", [[r[3]] for r in dept_formulas], sheet=SH))
ck("实发合计", set_formulas("F3", [[r[4]] for r in dept_formulas], sheet=SH))
ck("社保公司", set_formulas("G3", [[r[5]] for r in dept_formulas], sheet=SH))
ck("公积金个人",set_formulas("H3", [[r[6]] for r in dept_formulas], sheet=SH))
ck("公积金公司",set_formulas("I3", [[r[7]] for r in dept_formulas], sheet=SH))

last_dept = 2 + len(DEPTS_UNIQ)
ck("total label", set_values(f"A{last_dept+1}", [["合  计"]], sheet=SH))
for col_l in ["B","C","D","E","F","G","H","I"]:
    ck(f"total {col_l}", set_formulas(f"{col_l}{last_dept+1}",
        [[f"=SUM({col_l}3:{col_l}{last_dept})"]], sheet=SH))

ck("total bold", format_range(f"A{last_dept+1}:I{last_dept+1}", sheet=SH,
                              bold=True, bg_color="#E2EFDA"))

# Alternating row colors for dept table
for i in range(len(DEPTS_UNIQ)):
    r = i + 3
    color = "#F3E6FF" if i % 2 == 0 else "#FFFFFF"
    ck(f"row{r} color", format_range(f"A{r}:I{r}", sheet=SH, bg_color=color))

ck("number fmt CfHI", set_number_format(f"C3:I{last_dept+1}", "#,##0.00", sheet=SH))
ck("number fmt B",    set_number_format(f"B3:B{last_dept+1}", "0", sheet=SH))
ck("border",   format_range(f"A2:I{last_dept+1}", sheet=SH, border="all"))
ck("center A", format_range(f"A2:A{last_dept+1}", sheet=SH, h_align="center"))
ck("autofit",  autofit("A:I", sheet=SH))

# ── 保存 & 关闭 ────────────────────────────────────────────────────────────────
print("\n[ 保存 & 关闭 ]")
ck("save",  save_workbook())
ck("close", close_workbook(save=False))

# ── Summary ────────────────────────────────────────────────────────────────────
total  = len(results)
passed = sum(1 for _, ok in results if ok)
failed = [(lb, ok) for lb, ok in results if not ok]
print("\n" + "=" * 60)
print(f"  RESULT: {passed}/{total} passed")
if failed:
    print("  FAILED:")
    for lb, _ in failed:
        print(f"    - {lb}")
print("=" * 60)
print(f"\n  输出文件: {OUT}")
