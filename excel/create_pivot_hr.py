"""
在 HR社保财务统计.xlsx 中创建两张数据透视表 + 一张图表：
  透视表1 (薪酬透视): 薪酬统计 — 部门 × 应发合计/实发工资/个人所得税
  透视表2 (社保透视): 社保明细 — 部门 × 社保个人/公积金个人
  图表: 基于薪酬透视 的柱状图
"""
import sys, os, json, time
sys.path.insert(0, os.path.dirname(__file__))
os.environ.setdefault("EXCEL_SNAPSHOT_DIR", os.path.join(os.path.dirname(__file__), ".snapshots"))
os.environ.setdefault("EXCEL_MAX_SNAPSHOTS", "50")

import subprocess as _sp
_sp.run(["taskkill","/f","/im","et.exe"],  capture_output=True)
_sp.run(["taskkill","/f","/im","wps.exe"], capture_output=True)
time.sleep(4)

from excel_mcp.tools.file_tools   import open_workbook, save_workbook, close_workbook
from excel_mcp.tools.sheet_tools  import create_sheet, set_tab_color, delete_sheet
from excel_mcp.tools.pivot_tools  import (
    create_pivot_table, add_pivot_field,
    set_pivot_number_format, refresh_pivot,
    list_pivot_fields, list_pivot_tables,
)
from excel_mcp.tools.chart_tools  import (
    create_chart_from_pivot, set_chart_title, move_chart,
)
from excel_mcp.tools.range_tools  import (
    set_values, format_range, merge_cells,
    set_row_height, autofit, freeze_panes,
)

HR_FILE = r"C:\Users\zack_\AppData\Local\Temp\HR社保财务统计.xlsx"

results = []
def ck(label, raw, show_detail=False):
    d = json.loads(raw) if isinstance(raw, str) else raw
    ok = d.get("success", False)
    print(f"  {'OK' if ok else 'NG'}  {label}")
    if not ok:
        print(f"       ERR: {d.get('error','')[:120]}")
    elif show_detail:
        print(f"       {json.dumps(d, ensure_ascii=False)[:200]}")
    results.append((label, ok, d))
    return d

# ── 打开工作簿 ─────────────────────────────────────────────────────────────────
print("\n[ 0 ] 打开工作簿")
ck("open", open_workbook(HR_FILE))
time.sleep(1)

# 清理上次可能残留的"数据透视"工作表
import json as _json
_del = _json.loads(delete_sheet("数据透视"))
if _del.get("success"):
    print("  (已删除旧的数据透视工作表)")

ck("create 数据透视", create_sheet("数据透视"))
ck("tab color",       set_tab_color("数据透视", "#C00000"))

# ══════════════════════════════════════════════════════════════════════════════
# 透视表1: 薪酬统计 — 部门 × 应发合计 / 实发工资 / 个人所得税
# ══════════════════════════════════════════════════════════════════════════════
print("\n[ 1 ] 薪酬透视表")

# 页面标题
ck("title merge",  merge_cells("A1:H1", sheet="数据透视"))
ck("title text",   set_values("A1",
    [["XX 科技有限公司 — 2025年各部门薪酬 & 社保数据透视"]], sheet="数据透视"))
ck("title format", format_range("A1", sheet="数据透视",
    bold=True, font_size=13, h_align="center", v_align="center",
    bg_color="#C00000", font_color="#FFFFFF"))
ck("row1 height",  set_row_height(1, 34, sheet="数据透视"))

# 节标题
ck("sec1 label",  set_values("A2", [["▌ 薪酬汇总（按部门）"]], sheet="数据透视"))
ck("sec1 fmt",    format_range("A2:H2", sheet="数据透视",
    bold=True, font_color="#C00000", bg_color="#FFF2CC"))
ck("row2 height", set_row_height(2, 20, sheet="数据透视"))

# 数据源: 薪酬统计 A2:K17（表头+15行数据，不含第18行合计）
d = ck("create 薪酬透视", create_pivot_table(
    source_address="A2:K17",
    dest_cell="A3",
    name="薪酬透视",
    source_sheet="薪酬统计",
    dest_sheet="数据透视",
), show_detail=True)

pivot1_ok = d.get("success", False)
if pivot1_ok:
    # 查看实际字段名（WPS 有时会改名）
    fd = ck("fields", list_pivot_fields("薪酬透视", sheet="数据透视"), show_detail=True)
    field_names = {f["name"] for f in fd.get("fields", [])}
    print(f"       可用字段: {field_names}")

    ck("row 部门",      add_pivot_field("薪酬透视", "部门",     "row",  sheet="数据透视"))
    ck("data 应发合计", add_pivot_field("薪酬透视", "应发合计", "data", sheet="数据透视"))
    ck("data 实发工资", add_pivot_field("薪酬透视", "实发工资", "data", sheet="数据透视"))
    ck("data 个税",     add_pivot_field("薪酬透视", "个人所得税","data", sheet="数据透视"))

    # WPS 默认 SUM，不需要再设 Function；直接设数字格式
    ck("fmt 应发",  set_pivot_number_format("薪酬透视", "应发合计",   "#,##0.00", sheet="数据透视"))
    ck("fmt 实发",  set_pivot_number_format("薪酬透视", "实发工资",   "#,##0.00", sheet="数据透视"))
    ck("fmt 个税",  set_pivot_number_format("薪酬透视", "个人所得税", "#,##0.00", sheet="数据透视"))
    ck("refresh 1", refresh_pivot("薪酬透视", sheet="数据透视"))

# ══════════════════════════════════════════════════════════════════════════════
# 透视表2: 社保明细 — 部门 × 社保个人 / 公积金个人
# 放在第25行，安全避开透视表1（占 A3:C20）
# ══════════════════════════════════════════════════════════════════════════════
print("\n[ 2 ] 社保透视表")

ck("sec2 label",  set_values("A58", [["▌ 社保 & 公积金汇总（按部门）"]], sheet="数据透视"))
ck("sec2 fmt",    format_range("A58:H58", sheet="数据透视",
    bold=True, font_color="#C00000", bg_color="#FFF2CC"))
ck("row58 height", set_row_height(58, 20, sheet="数据透视"))

# 数据源: 社保明细 A3:P18（第3行表头，第4-18行15条数据，不含合计行19）
# 放在第59行起，安全避开薪酬透视（WPS 刷新后可扩展至A3:A54）
d2 = ck("create 社保透视", create_pivot_table(
    source_address="A3:P18",
    dest_cell="A59",
    name="社保透视",
    source_sheet="社保明细",
    dest_sheet="数据透视",
), show_detail=True)

pivot2_ok = d2.get("success", False)
if pivot2_ok:
    fd2 = ck("fields 社保", list_pivot_fields("社保透视", sheet="数据透视"), show_detail=True)
    field_names2 = {f["name"] for f in fd2.get("fields", [])}
    print(f"       可用字段: {field_names2}")

    ck("row 部门",       add_pivot_field("社保透视", "部门",           "row",  sheet="数据透视"))
    ck("data 社保个人",  add_pivot_field("社保透视", "社保合计(个人)", "data", sheet="数据透视"))
    ck("data 公积金个人",add_pivot_field("社保透视", "公积金合计(个人)","data", sheet="数据透视"))
    ck("fmt 社保",  set_pivot_number_format("社保透视","社保合计(个人)",  "#,##0.00", sheet="数据透视"))
    ck("fmt 公积金",set_pivot_number_format("社保透视","公积金合计(个人)","#,##0.00", sheet="数据透视"))
    ck("refresh 2", refresh_pivot("社保透视", sheet="数据透视"))

# ══════════════════════════════════════════════════════════════════════════════
# 图表: 基于薪酬透视表的柱状图（放在透视表右侧）
# ══════════════════════════════════════════════════════════════════════════════
print("\n[ 3 ] 创建图表")

if pivot1_ok:
    dc = ck("chart from pivot", create_chart_from_pivot(
        pivot_name="薪酬透视",
        chart_type="column_clustered",
        title="各部门薪酬汇总",
        left=380, top=40,
        width=420, height=280,
        sheet="数据透视",
    ), show_detail=True)

# ── 收尾 ──────────────────────────────────────────────────────────────────────
ck("autofit",  autofit("A:H", sheet="数据透视"))
ck("freeze",   freeze_panes(2, 0, sheet="数据透视"))

# ── 列出所有透视表确认 ─────────────────────────────────────────────────────────
ck("list all pivots", list_pivot_tables(sheet="数据透视"), show_detail=True)

# ── 保存 & 关闭 ────────────────────────────────────────────────────────────────
print("\n[ 4 ] 保存 & 关闭")
# SaveAs 到新文件名以避开 WPS 对当前打开文件的锁定
PIVOT_FILE = HR_FILE.replace(".xlsx", "_含透视表.xlsx")
ck("save as new file", save_workbook(save_as=PIVOT_FILE))
ck("close",            close_workbook(save=False))

# ── Summary ────────────────────────────────────────────────────────────────────
total  = len(results)
passed = sum(1 for _, ok, _ in results if ok)
failed = [(lb, d) for lb, ok, d in results if not ok]
print("\n" + "=" * 60)
print(f"  RESULT: {passed}/{total} passed")
if failed:
    print("  FAILED:")
    for lb, d in failed:
        print(f"    - {lb}: {d.get('error','')[:80]}")
print("=" * 60)

# 验证文件是否正确包含"数据透视"工作表
import openpyxl as _ox, warnings as _w
_w.filterwarnings("ignore")
try:
    _wb = _ox.load_workbook(PIVOT_FILE)
    print(f"\n  文件: {PIVOT_FILE}")
    print(f"  工作表: {_wb.sheetnames}")
except Exception as _e:
    print(f"\n  验证失败: {_e}")
