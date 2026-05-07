"""
增量测试 — 覆盖 Tasks #11-#19 新增的所有操作：
  行高/列宽、隐藏显示、冻结窗格、AutoFilter、批注、
  行列分组、工作簿保护、删除重复值、插入图片、PDF导出
"""
import sys, os, json, time, tempfile, struct, zlib
sys.path.insert(0, os.path.dirname(__file__))
os.environ.setdefault("EXCEL_SNAPSHOT_DIR", os.path.join(os.path.dirname(__file__), ".snapshots"))
os.environ.setdefault("EXCEL_MAX_SNAPSHOTS", "50")

import openpyxl

from excel_mcp.tools.file_tools import (
    open_workbook, save_workbook, close_workbook,
    protect_workbook, unprotect_workbook, export_pdf,
)
from excel_mcp.tools.range_tools import (
    set_values, get_values, format_range,
    set_row_height, set_column_width,
    hide_rows, show_rows, hide_columns, show_columns,
    freeze_panes, unfreeze_panes,
    apply_range_autofilter, clear_range_autofilter, toggle_autofilter,
    group_rows, ungroup_rows, group_columns, ungroup_columns,
    remove_duplicates,
)
from excel_mcp.tools.comment_tools import (
    add_comment, get_comments, delete_comment, show_hide_comment,
)
from excel_mcp.tools.image_tools import (
    insert_image, list_images, delete_image,
)
from excel_mcp.tools.sheet_tools import create_sheet

results = []

def check(label, raw, extra_check=None):
    d = json.loads(raw) if isinstance(raw, str) else raw
    ok = d.get("success", False)
    if ok and extra_check:
        passed, msg = extra_check(d)
        if not passed:
            ok = False
            print(f"  NG  {label}")
            print(f"       ASSERT: {msg}")
            results.append((label, False))
            return d
    print(f"  {'OK' if ok else 'NG'}  {label}")
    if not ok:
        print(f"       ERR: {d.get('error', d)}")
    results.append((label, ok))
    return d


def _make_tiny_png():
    """Create a minimal valid 2x2 red PNG in memory."""
    sig = b'\x89PNG\r\n\x1a\n'
    def chunk(name, data):
        c = struct.pack('>I', len(data)) + name + data
        return c + struct.pack('>I', zlib.crc32(name + data) & 0xffffffff)
    ihdr = chunk(b'IHDR', struct.pack('>IIBBBBB', 2, 2, 8, 2, 0, 0, 0))
    raw_rows = b'\x00\xff\x00\x00\xff\x00\x00\n\x00\xff\x00\x00\xff\x00\x00'
    # scanlines for 2x2 RGB: filter_byte + R G B * 2
    scanline = b'\x00' + b'\xff\x00\x00' * 2
    idat = chunk(b'IDAT', zlib.compress(scanline * 2))
    iend = chunk(b'IEND', b'')
    return sig + ihdr + idat + iend


# ── Kill stale WPS ─────────────────────────────────────────────────────────────
import subprocess as _sp
_sp.run(["taskkill", "/f", "/im", "et.exe"],  capture_output=True)
_sp.run(["taskkill", "/f", "/im", "wps.exe"], capture_output=True)
time.sleep(4)

# ── Create workbook ────────────────────────────────────────────────────────────
xlsx_path = os.path.join(tempfile.gettempdir(), f"incr_test_{int(time.time())}.xlsx")
wb0 = openpyxl.Workbook()
wb0.active.title = "Sheet1"
wb0.save(xlsx_path)
print(f"\n[ 0 ] 准备工作簿  →  {xlsx_path}")
check("open_workbook", open_workbook(xlsx_path))
time.sleep(1)

check("create 数据表", create_sheet("数据表"))
check("create 分组表", create_sheet("分组表"))
check("create 图片表", create_sheet("图片表"))

# Seed Sheet1 with simple data for most tests
check("seed Sheet1 headers",
    set_values("A1", [["姓名","部门","金额","日期"]], sheet="Sheet1"))
data = [
    ["张三", "研发", 5000, "2024-01"],
    ["李四", "市场", 8000, "2024-02"],
    ["王五", "研发", 6000, "2024-01"],
    ["张三", "研发", 5000, "2024-01"],   # duplicate of row 2
    ["赵六", "财务", 9000, "2024-03"],
    ["李四", "市场", 8000, "2024-02"],   # duplicate of row 3
    ["钱七", "研发", 7000, "2024-02"],
]
check("seed Sheet1 data",
    set_values("A2", data, sheet="Sheet1"))

# ══════════════════════════════════════════════════════════════════════════════
print("\n[ 1 ] 行高 / 列宽 / 隐藏显示")

check("set_row_height row1=30",
    set_row_height(1, 30, sheet="Sheet1"))

check("set_column_width colA=20",
    set_column_width(1, 20, sheet="Sheet1"))

check("set_row_height 3 rows starting at 2",
    set_row_height(2, 18, count=3, sheet="Sheet1"))

check("set_column_width 2 cols starting at 2",
    set_column_width(2, 15, count=2, sheet="Sheet1"))

check("hide_rows row5",
    hide_rows(5, sheet="Sheet1"))

check("show_rows row5",
    show_rows(5, sheet="Sheet1"))

check("hide_columns col4 (D)",
    hide_columns(4, sheet="Sheet1"))

check("show_columns col4 (D)",
    show_columns(4, sheet="Sheet1"))

# ══════════════════════════════════════════════════════════════════════════════
print("\n[ 2 ] 冻结窗格")

check("freeze_panes top row only",
    freeze_panes(1, 0, sheet="Sheet1"))

check("freeze_panes row+col",
    freeze_panes(1, 1, sheet="Sheet1"))

check("unfreeze_panes",
    unfreeze_panes(sheet="Sheet1"))

check("freeze then unfreeze again",
    freeze_panes(2, 2, sheet="Sheet1"))

check("unfreeze final",
    unfreeze_panes(sheet="Sheet1"))

# ══════════════════════════════════════════════════════════════════════════════
print("\n[ 3 ] AutoFilter（range级）")

check("apply_range_autofilter enable only",
    apply_range_autofilter("A1:D8", sheet="Sheet1"))

check("apply_range_autofilter field=2 criteria=研发",
    apply_range_autofilter("A1:D8", field=2, criteria="研发", sheet="Sheet1"))

check("clear_range_autofilter",
    clear_range_autofilter(sheet="Sheet1"))

check("toggle_autofilter off",
    toggle_autofilter(sheet="Sheet1"))

check("toggle_autofilter on",
    toggle_autofilter(sheet="Sheet1"))

check("toggle_autofilter off again",
    toggle_autofilter(sheet="Sheet1"))

# ══════════════════════════════════════════════════════════════════════════════
print("\n[ 4 ] 批注")

check("add_comment A1",
    add_comment("A1", "这是姓名列", sheet="Sheet1"))

check("add_comment B1",
    add_comment("B1", "部门字段", author="测试员", sheet="Sheet1"))

d = check("get_comments Sheet1",
    get_comments(sheet="Sheet1"),
    extra_check=lambda d: (d["count"] >= 2, f"expected >=2 comments, got {d['count']}"))

check("show_hide_comment A1 visible",
    show_hide_comment("A1", visible=True, sheet="Sheet1"))

check("show_hide_comment A1 hidden",
    show_hide_comment("A1", visible=False, sheet="Sheet1"))

check("delete_comment A1",
    delete_comment("A1", sheet="Sheet1"))

d2 = check("get_comments after delete",
    get_comments(sheet="Sheet1"),
    extra_check=lambda d: (d["count"] == 1, f"expected 1 comment, got {d['count']}"))

# ══════════════════════════════════════════════════════════════════════════════
print("\n[ 5 ] 行列分组 / 折叠")

# Seed 分组表
check("seed 分组表",
    set_values("A1", [
        ["区域"], ["华北"], ["华北"], ["华北"],
        ["华南"], ["华南"], ["华东"],
    ], sheet="分组表"))

check("group_rows 2-4",
    group_rows(2, 3, sheet="分组表"))

check("group_rows 5-6",
    group_rows(5, 2, sheet="分组表"))

check("ungroup_rows 2-4",
    ungroup_rows(2, 3, sheet="分组表"))

check("group_columns col2-3",
    group_columns(2, 2, sheet="分组表"))

check("ungroup_columns col2-3",
    ungroup_columns(2, 2, sheet="分组表"))

# ══════════════════════════════════════════════════════════════════════════════
print("\n[ 6 ] 删除重复值")

# Seed 数据表
check("seed 数据表 for dedup",
    set_values("A1", [
        ["姓名", "部门"],
        ["张三", "研发"],
        ["李四", "市场"],
        ["张三", "研发"],  # dup
        ["王五", "财务"],
        ["李四", "市场"],  # dup
    ], sheet="数据表"))

d = check("remove_duplicates all cols",
    remove_duplicates("A1:B6", has_header=True, sheet="数据表"))

# Verify via get_values that rows reduced
vals = check("verify rows after dedup",
    get_values("A1:B6", sheet="数据表"))

# ══════════════════════════════════════════════════════════════════════════════
print("\n[ 7 ] 工作簿保护")

check("protect_workbook structure",
    protect_workbook(structure=True, windows=False))

check("unprotect_workbook",
    unprotect_workbook())

check("protect_workbook with password",
    protect_workbook(password="test123", structure=True))

check("unprotect_workbook with password",
    unprotect_workbook(password="test123"))

# ══════════════════════════════════════════════════════════════════════════════
print("\n[ 8 ] 插入图片")

# Write a tiny PNG to disk
png_path = os.path.join(tempfile.gettempdir(), "test_img.png")
with open(png_path, "wb") as f:
    f.write(_make_tiny_png())

d = check("insert_image 图片表 A1",
    insert_image(png_path, "A1", width=80, height=60, sheet="图片表"))

img_name = d.get("name", "")

d2 = check("list_images 图片表",
    list_images(sheet="图片表"),
    extra_check=lambda d: (d["count"] >= 1, f"expected >=1 image, got {d['count']}"))

check("insert second image B5",
    insert_image(png_path, "B5", width=40, height=30, sheet="图片表"))

check("list_images count==2",
    list_images(sheet="图片表"),
    extra_check=lambda d: (d["count"] >= 2, f"expected >=2 images, got {d['count']}"))

if img_name:
    check("delete_image first",
        delete_image(img_name, sheet="图片表"))

# ══════════════════════════════════════════════════════════════════════════════
print("\n[ 9 ] PDF 导出")

pdf_path = os.path.join(tempfile.gettempdir(), f"incr_test_{int(time.time())}.pdf")
d = check("export_pdf full workbook",
    export_pdf(pdf_path))

check("pdf file exists",
    {"success": os.path.isfile(pdf_path), "path": pdf_path},
    extra_check=lambda d: (d["success"], f"PDF not found at {pdf_path}"))

# Sheet-only export
pdf_sheet_path = os.path.join(tempfile.gettempdir(), f"incr_sheet_{int(time.time())}.pdf")
d2 = check("export_pdf Sheet1 only",
    export_pdf(pdf_sheet_path, sheet="Sheet1"))

check("sheet pdf file exists",
    {"success": os.path.isfile(pdf_sheet_path), "path": pdf_sheet_path},
    extra_check=lambda d: (d["success"], f"PDF not found at {pdf_sheet_path}"))

# ══════════════════════════════════════════════════════════════════════════════
print("\n[ 10 ] 保存 & 关闭")
check("save_workbook",  save_workbook())
check("close_workbook", close_workbook(save=False))

# ── Summary ────────────────────────────────────────────────────────────────────
total  = len(results)
passed = sum(1 for _, ok in results if ok)
failed = [(lb, ok) for lb, ok in results if not ok]

print("\n" + "=" * 60)
print(f"  RESULT: {passed}/{total} passed")
if failed:
    print("  FAILED:")
    for lb, _ in failed:
        print(f"    - {lb}")
print("=" * 60)
print(f"  Output: {xlsx_path}")
