"""End-to-end test for Excel MCP tools — runs directly against excel_mcp package."""
import sys, os, json, time, tempfile
sys.path.insert(0, os.path.dirname(__file__))
os.environ.setdefault("EXCEL_SNAPSHOT_DIR", os.path.join(os.path.dirname(__file__), ".snapshots"))
os.environ.setdefault("EXCEL_MAX_SNAPSHOTS", "20")

import openpyxl
from excel_mcp.tools.file_tools   import list_sessions, open_workbook, save_workbook, close_workbook, get_workbook_info
from excel_mcp.tools.range_tools  import get_values, set_values, set_formulas, format_range, set_number_format, get_used_range
from excel_mcp.tools.sheet_tools  import list_sheets, create_sheet, rename_sheet, set_tab_color
from excel_mcp.tools.table_tools  import create_table, get_table_data, append_rows_to_table, apply_table_style
from excel_mcp.tools.chart_tools  import create_chart, list_charts
from excel_mcp.tools.pivot_tools  import create_pivot_table, add_pivot_field, refresh_pivot, list_pivot_tables
from excel_mcp.tools.undo_tools   import list_undo_snapshots, undo_last
from excel_mcp.tools.calc_tools   import get_calculation_mode

PASS = "OK"
FAIL = "NG"
results = []

def check(label, raw):
    d = json.loads(raw) if isinstance(raw, str) else raw
    ok = d.get("success", False)
    mark = PASS if ok else FAIL
    print(f"  {mark}  {label}")
    if not ok:
        print(f"       ERROR: {d.get('error', d)}")
    results.append((label, ok))
    return d

# ── 1. 创建测试 xlsx ────────────────────────────────────────────────────────
print("\n[ 1 ] 准备测试文件")
xlsx_path = os.path.join(tempfile.gettempdir(), f"excel_mcp_test_{int(time.time())}.xlsx")
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Sales"
ws.append(["Region", "Product", "Q1", "Q2", "Q3", "Q4"])
data = [
    ["East",  "Widget A", 12000, 15000, 13500, 17000],
    ["East",  "Widget B",  8000,  9200,  8800, 11000],
    ["West",  "Widget A", 10500, 11000, 12000, 14500],
    ["West",  "Widget B",  7200,  8100,  9000, 10200],
    ["North", "Widget A",  9000, 10000, 11500, 13000],
    ["North", "Widget B",  6000,  7500,  8000,  9500],
    ["South", "Widget A", 11000, 12000, 13000, 15000],
    ["South", "Widget B",  7800,  8500,  9200, 10800],
]
for row in data:
    ws.append(row)
wb.save(xlsx_path)
print(f"  {PASS}  Created: {xlsx_path}")

# ── 2. 打开工作簿 ────────────────────────────────────────────────────────────
print("\n[ 2 ] 文件 / 会话")
check("open_workbook", open_workbook(xlsx_path))
time.sleep(1)
check("list_sessions", list_sessions())
check("get_workbook_info", get_workbook_info())

# ── 3. 范围读写 ──────────────────────────────────────────────────────────────
print("\n[ 3 ] 范围读写")
check("get_used_range", get_used_range())
r = check("get_values A1:F3", get_values("A1:F3"))
print(f"       rows={len(r.get('values',[]))}, cols={len(r.get('values',[[]])[0])}")
check("set_values  G1=Total", set_values("G1", [["Total"]]))
check("set_formulas G2:G9", set_formulas("G2", [[f"=SUM(C{i}:F{i})"] for i in range(2, 10)]))

# ── 4. 格式化 ────────────────────────────────────────────────────────────────
print("\n[ 4 ] 格式化")
check("format header row",   format_range("A1:G1", bold=True, bg_color="#1F4E79", font_color="#FFFFFF", h_align="center"))
check("number format C:G",   set_number_format("C2:G9", "#,##0"))
check("format even rows",    format_range("A2:G9", bg_color="#EBF3FB"))

# ── 5. 表格 ──────────────────────────────────────────────────────────────────
print("\n[ 5 ] Excel 表格")
check("create_table SalesData", create_table("A1:G9", "SalesData"))
check("apply_table_style",      apply_table_style("SalesData", "TableStyleMedium9"))
r = check("get_table_data",     get_table_data("SalesData"))
print(f"       rows={r.get('total_rows')}, cols={r.get('cols')}")
check("append_rows_to_table",   append_rows_to_table("SalesData", [["South","Widget C",5000,6000,7000,8000,26000]]))

# ── 6. 工作表操作 ────────────────────────────────────────────────────────────
print("\n[ 6 ] 工作表")
check("list_sheets",     list_sheets())
check("create_sheet Pivot", create_sheet("Pivot"))
check("create_sheet Charts", create_sheet("Charts"))
check("set_tab_color Sales", set_tab_color("Sales", "#1F4E79"))
check("set_tab_color Pivot", set_tab_color("Pivot", "#375623"))

# ── 7. 透视表 ────────────────────────────────────────────────────────────────
print("\n[ 7 ] 透视表")
check("create_pivot_table", create_pivot_table("A1:G9", "B3", "RegionPivot", source_sheet="Sales", dest_sheet="Pivot"))
time.sleep(1)
check("add_pivot ROW Region",   add_pivot_field("RegionPivot", "Region",  "row",    sheet="Pivot"))
check("add_pivot COL Product",  add_pivot_field("RegionPivot", "Product", "column", sheet="Pivot"))
check("add_pivot VAL Total",    add_pivot_field("RegionPivot", "Total",   "value",  sheet="Pivot"))
check("refresh_pivot",          refresh_pivot("RegionPivot", "Pivot"))
r = check("list_pivot_tables",  list_pivot_tables("Pivot"))
print(f"       pivots={r.get('count')}")

# ── 8. 图表 ──────────────────────────────────────────────────────────────────
print("\n[ 8 ] 图表")
check("create_chart bar",    create_chart("A1:G9", chart_type="bar_clustered", title="Sales by Region & Product", sheet="Sales"))
time.sleep(0.5)
check("create_chart column", create_chart("C1:G9", chart_type="column_clustered", title="Quarterly Trend", sheet="Sales"))
r = check("list_charts",     list_charts("Charts"))
print(f"       charts={r.get('count')}")

# ── 9. 计算 & 快照 ───────────────────────────────────────────────────────────
print("\n[ 9 ] 计算模式 & 快照")
check("get_calculation_mode", get_calculation_mode())
r = check("list_undo_snapshots", list_undo_snapshots())
print(f"       snapshots={r.get('count')}")

# ── 10. 保存 ─────────────────────────────────────────────────────────────────
print("\n[ 10 ] 保存 & 关闭")
check("save_workbook",  save_workbook())
check("close_workbook", close_workbook())

# ── 汇总 ─────────────────────────────────────────────────────────────────────
passed = sum(1 for _, ok in results if ok)
total  = len(results)
print(f"\n{'='*50}")
print(f"  RESULT: {passed}/{total} passed")
if passed < total:
    print("  FAILED:")
    for label, ok in results:
        if not ok:
            print(f"    - {label}")
print(f"{'='*50}")
print(f"  Test file: {xlsx_path}")
