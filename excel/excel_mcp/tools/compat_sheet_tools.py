"""Cross-platform sheet management tools — openpyxl backend."""
import json
from excel_mcp.core.openpyxl_backend import OpenpyxlSession, OpenpyxlSnapshotUndo, parse_color


def _ctx(workbook, sheet=None):
    name, session = OpenpyxlSession.get(workbook)
    wb = session["wb"]
    ws = OpenpyxlSession.get_sheet(wb, sheet) if sheet else None
    return name, session, wb, ws


def list_sheets(workbook: str = None) -> str:
    """List all worksheets in a workbook."""
    try:
        name, session, wb, _ = _ctx(workbook)
        sheets = [
            {
                "name": ws.title,
                "index": idx + 1,
                "visible": ws.sheet_state == "visible",
                "very_hidden": ws.sheet_state == "veryHidden",
                "tab_color": (
                    f"#{ws.sheet_properties.tabColor.rgb[2:]}"
                    if ws.sheet_properties.tabColor else None
                ),
            }
            for idx, ws in enumerate(wb.worksheets)
        ]
        return json.dumps({"success": True, "sheets": sheets, "count": len(sheets)})
    except Exception as e:
        return json.dumps({"success": False, "error": str(e)})


def create_sheet(name: str, position: int = None, workbook: str = None) -> str:
    """Create a new worksheet. *position* is 0-based (None = last)."""
    try:
        wb_name, session, wb, _ = _ctx(workbook)
        OpenpyxlSnapshotUndo.push(wb_name, session)
        ws = wb.create_sheet(title=name, index=position)
        wb.save(session["path"])
        return json.dumps({"success": True, "name": ws.title,
                           "index": wb.sheetnames.index(ws.title) + 1})
    except Exception as e:
        return json.dumps({"success": False, "error": str(e)})


def rename_sheet(old_name: str, new_name: str, workbook: str = None) -> str:
    """Rename a worksheet."""
    try:
        wb_name, session, wb, _ = _ctx(workbook)
        OpenpyxlSnapshotUndo.push(wb_name, session)
        ws = OpenpyxlSession.get_sheet(wb, old_name)
        ws.title = new_name
        wb.save(session["path"])
        return json.dumps({"success": True, "old_name": old_name, "new_name": new_name})
    except Exception as e:
        return json.dumps({"success": False, "error": str(e)})


def copy_sheet(sheet: str, new_name: str = None, position: int = None,
               workbook: str = None) -> str:
    """Copy a worksheet within the same workbook."""
    try:
        wb_name, session, wb, _ = _ctx(workbook)
        OpenpyxlSnapshotUndo.push(wb_name, session)
        ws_src = OpenpyxlSession.get_sheet(wb, sheet)
        new_ws = wb.copy_worksheet(ws_src)
        if new_name:
            new_ws.title = new_name
        if position is not None:
            sheets = wb._sheets
            sheets.remove(new_ws)
            sheets.insert(position, new_ws)
        wb.save(session["path"])
        return json.dumps({"success": True, "original": sheet, "copy": new_ws.title,
                           "index": wb.sheetnames.index(new_ws.title) + 1})
    except Exception as e:
        return json.dumps({"success": False, "error": str(e)})


def move_sheet(sheet: str, position: int, workbook: str = None) -> str:
    """Move a worksheet to a new 0-based position."""
    try:
        wb_name, session, wb, _ = _ctx(workbook)
        OpenpyxlSnapshotUndo.push(wb_name, session)
        ws = OpenpyxlSession.get_sheet(wb, sheet)
        wb.move_sheet(ws, offset=position - wb.sheetnames.index(ws.title))
        wb.save(session["path"])
        return json.dumps({"success": True, "sheet": sheet,
                           "new_index": wb.sheetnames.index(ws.title) + 1})
    except Exception as e:
        return json.dumps({"success": False, "error": str(e)})


def delete_sheet(sheet: str, workbook: str = None) -> str:
    """Delete a worksheet (snapshot taken first)."""
    try:
        wb_name, session, wb, _ = _ctx(workbook)
        OpenpyxlSnapshotUndo.push(wb_name, session)
        ws = OpenpyxlSession.get_sheet(wb, sheet)
        del wb[ws.title]
        wb.save(session["path"])
        return json.dumps({"success": True, "deleted": sheet})
    except Exception as e:
        return json.dumps({"success": False, "error": str(e)})


def set_tab_color(sheet: str, color: str = None, workbook: str = None) -> str:
    """Set sheet tab color (#RRGGBB). Pass color=None to clear."""
    try:
        wb_name, session, wb, _ = _ctx(workbook)
        ws = OpenpyxlSession.get_sheet(wb, sheet)
        if color is None:
            ws.sheet_properties.tabColor = None
        else:
            from openpyxl.styles.colors import Color
            ws.sheet_properties.tabColor = Color(rgb=parse_color(color))
        wb.save(session["path"])
        return json.dumps({"success": True, "sheet": sheet, "color": color})
    except Exception as e:
        return json.dumps({"success": False, "error": str(e)})


def set_sheet_visibility(sheet: str, visibility: str = "visible",
                         workbook: str = None) -> str:
    """Set sheet visibility. visibility: visible | hidden | very_hidden"""
    try:
        wb_name, session, wb, _ = _ctx(workbook)
        ws = OpenpyxlSession.get_sheet(wb, sheet)
        state_map = {
            "visible": "visible",
            "hidden": "hidden",
            "very_hidden": "veryHidden",
        }
        ws.sheet_state = state_map.get(visibility, "visible")
        wb.save(session["path"])
        return json.dumps({"success": True, "sheet": sheet, "visibility": visibility})
    except Exception as e:
        return json.dumps({"success": False, "error": str(e)})


def copy_sheet_to_file(sheet: str, dest_path: str, dest_sheet_name: str = None,
                       workbook: str = None) -> str:
    """Copy a worksheet into another workbook file."""
    try:
        import shutil
        from pathlib import Path
        import openpyxl

        wb_name, session, wb, _ = _ctx(workbook)
        ws_src = OpenpyxlSession.get_sheet(wb, sheet)
        dest_abs = str(Path(dest_path).resolve())

        if Path(dest_abs).exists():
            dest_wb = openpyxl.load_workbook(dest_abs)
        else:
            dest_wb = openpyxl.Workbook()
            # Remove default empty sheet
            if "Sheet" in dest_wb.sheetnames:
                del dest_wb["Sheet"]

        new_ws = dest_wb.copy_worksheet(ws_src)  # Note: only works within same wb object
        # Workaround: save src to temp, copy sheet data manually
        new_ws = dest_wb.create_sheet(title=dest_sheet_name or ws_src.title)
        for row in ws_src.iter_rows():
            for cell in row:
                new_ws.cell(row=cell.row, column=cell.column, value=cell.value)

        dest_wb.save(dest_abs)
        return json.dumps({"success": True, "sheet": new_ws.title, "dest": dest_abs})
    except Exception as e:
        return json.dumps({"success": False, "error": str(e)})


def protect_sheet(sheet: str, password: str = None, workbook: str = None) -> str:
    """Protect a worksheet with optional password."""
    try:
        wb_name, session, wb, _ = _ctx(workbook)
        ws = OpenpyxlSession.get_sheet(wb, sheet)
        ws.protection.sheet = True
        if password:
            ws.protection.password = password
        wb.save(session["path"])
        return json.dumps({"success": True, "sheet": sheet, "protected": True})
    except Exception as e:
        return json.dumps({"success": False, "error": str(e)})


def unprotect_sheet(sheet: str, password: str = None, workbook: str = None) -> str:
    """Remove worksheet protection."""
    try:
        wb_name, session, wb, _ = _ctx(workbook)
        ws = OpenpyxlSession.get_sheet(wb, sheet)
        ws.protection.sheet = False
        wb.save(session["path"])
        return json.dumps({"success": True, "sheet": sheet, "protected": False})
    except Exception as e:
        return json.dumps({"success": False, "error": str(e)})
