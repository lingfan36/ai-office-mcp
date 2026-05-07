"""Cross-platform range tools — openpyxl backend (no Office required)."""
import json
import re
from typing import Any

from openpyxl.styles import (
    Font, PatternFill, Border, Side, Alignment, numbers
)
from openpyxl.styles.numbers import FORMAT_NUMBER, FORMAT_NUMBER_00
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.utils.cell import range_boundaries
from openpyxl.worksheet.datavalidation import DataValidation

from excel_mcp.core.openpyxl_backend import (
    OpenpyxlSession, OpenpyxlSnapshotUndo, parse_color,
)


# ── helpers ──────────────────────────────────────────────────────────────────

def _get_context(workbook, sheet):
    name, session = OpenpyxlSession.get(workbook)
    wb = session["wb"]
    ws = OpenpyxlSession.get_sheet(wb, sheet)
    return name, session, wb, ws


def _iter_range(ws, address: str):
    """Yield individual cells in an address (handles single cell + range)."""
    for row in ws[address]:
        if hasattr(row, '__iter__'):
            yield from row
        else:
            yield row


def _side(style: str, color: str = None) -> Side:
    s = Side(border_style=style or "thin")
    if color:
        from openpyxl.styles.colors import Color
        s.color = Color(rgb=parse_color(color))
    return s


# ── values / formulas ────────────────────────────────────────────────────────

def get_values(address: str, sheet: str = None, workbook: str = None) -> str:
    """Read cell values from a range. Returns 2-D list."""
    try:
        _, _, _, ws = _get_context(workbook, sheet)
        cells = ws[address]
        if not hasattr(cells, '__iter__') or isinstance(cells[0] if cells else None, tuple):
            # Multi-row range
            if not hasattr(cells, '__iter__'):
                data = [[cells.value]]
            else:
                data = [[c.value for c in row] for row in cells]
        else:
            # Single row
            data = [[c.value for c in cells]]
        return json.dumps({"success": True, "address": address, "values": data,
                           "rows": len(data), "cols": len(data[0]) if data else 0})
    except Exception as e:
        return json.dumps({"success": False, "error": str(e)})


def set_values(address: str, values: list[list[Any]], sheet: str = None,
               workbook: str = None, snapshot: bool = True) -> str:
    """Write a 2-D list of values into a range."""
    try:
        name, session, wb, ws = _get_context(workbook, sheet)
        if snapshot:
            OpenpyxlSnapshotUndo.push(name, session)
        min_col, min_row, _, _ = range_boundaries(address)
        for r, row in enumerate(values):
            for c, val in enumerate(row):
                ws.cell(row=min_row + r, column=min_col + c, value=val)
        wb.save(session["path"])
        return json.dumps({"success": True, "address": address,
                           "rows_written": len(values),
                           "cols_written": max(len(r) for r in values) if values else 0})
    except Exception as e:
        return json.dumps({"success": False, "error": str(e)})


def get_formulas(address: str, sheet: str = None, workbook: str = None) -> str:
    """Read cell formulas from a range. Returns 2-D list (formula or value)."""
    try:
        _, _, _, ws = _get_context(workbook, sheet)
        cells = ws[address]
        if not hasattr(cells[0] if cells else None, '__iter__'):
            data = [[c.value for c in cells]]
        else:
            data = [[c.value for c in row] for row in cells]
        return json.dumps({"success": True, "address": address, "formulas": data})
    except Exception as e:
        return json.dumps({"success": False, "error": str(e)})


def set_formulas(address: str, formulas: list[list[str]], sheet: str = None,
                 workbook: str = None, snapshot: bool = True) -> str:
    """Write formulas into a range."""
    return set_values(address, formulas, sheet=sheet, workbook=workbook, snapshot=snapshot)


# ── structure ─────────────────────────────────────────────────────────────────

def clear_range(address: str, sheet: str = None, workbook: str = None) -> str:
    """Clear all values in a range (preserves formatting)."""
    try:
        name, session, wb, ws = _get_context(workbook, sheet)
        OpenpyxlSnapshotUndo.push(name, session)
        for cell in _iter_range(ws, address):
            cell.value = None
        wb.save(session["path"])
        return json.dumps({"success": True, "cleared": address})
    except Exception as e:
        return json.dumps({"success": False, "error": str(e)})


def insert_rows(row: int, count: int = 1, sheet: str = None, workbook: str = None) -> str:
    """Insert *count* rows before *row* (1-based)."""
    try:
        name, session, wb, ws = _get_context(workbook, sheet)
        OpenpyxlSnapshotUndo.push(name, session)
        ws.insert_rows(row, count)
        wb.save(session["path"])
        return json.dumps({"success": True, "inserted_before": row, "count": count})
    except Exception as e:
        return json.dumps({"success": False, "error": str(e)})


def delete_rows(row: int, count: int = 1, sheet: str = None, workbook: str = None) -> str:
    """Delete *count* rows starting at *row* (1-based)."""
    try:
        name, session, wb, ws = _get_context(workbook, sheet)
        OpenpyxlSnapshotUndo.push(name, session)
        ws.delete_rows(row, count)
        wb.save(session["path"])
        return json.dumps({"success": True, "deleted_from": row, "count": count})
    except Exception as e:
        return json.dumps({"success": False, "error": str(e)})


def insert_columns(col: int, count: int = 1, sheet: str = None, workbook: str = None) -> str:
    """Insert *count* columns before *col* (1-based)."""
    try:
        name, session, wb, ws = _get_context(workbook, sheet)
        OpenpyxlSnapshotUndo.push(name, session)
        ws.insert_cols(col, count)
        wb.save(session["path"])
        return json.dumps({"success": True, "inserted_before_col": col, "count": count})
    except Exception as e:
        return json.dumps({"success": False, "error": str(e)})


def delete_columns(col: int, count: int = 1, sheet: str = None, workbook: str = None) -> str:
    """Delete *count* columns starting at *col* (1-based)."""
    try:
        name, session, wb, ws = _get_context(workbook, sheet)
        OpenpyxlSnapshotUndo.push(name, session)
        ws.delete_cols(col, count)
        wb.save(session["path"])
        return json.dumps({"success": True, "deleted_from_col": col, "count": count})
    except Exception as e:
        return json.dumps({"success": False, "error": str(e)})


def merge_cells(address: str, sheet: str = None, workbook: str = None) -> str:
    """Merge cells in a range."""
    try:
        name, session, wb, ws = _get_context(workbook, sheet)
        OpenpyxlSnapshotUndo.push(name, session)
        ws.merge_cells(address)
        wb.save(session["path"])
        return json.dumps({"success": True, "merged": address})
    except Exception as e:
        return json.dumps({"success": False, "error": str(e)})


def unmerge_cells(address: str, sheet: str = None, workbook: str = None) -> str:
    """Unmerge cells in a range."""
    try:
        name, session, wb, ws = _get_context(workbook, sheet)
        OpenpyxlSnapshotUndo.push(name, session)
        ws.unmerge_cells(address)
        wb.save(session["path"])
        return json.dumps({"success": True, "unmerged": address})
    except Exception as e:
        return json.dumps({"success": False, "error": str(e)})


def copy_range(src_address: str, dest_address: str, sheet: str = None,
               dest_sheet: str = None, workbook: str = None) -> str:
    """Copy values (and basic formatting) from one range to another."""
    try:
        name, session, wb, ws_src = _get_context(workbook, sheet)
        ws_dst = OpenpyxlSession.get_sheet(wb, dest_sheet) if dest_sheet else ws_src
        OpenpyxlSnapshotUndo.push(name, session)

        src_cells = ws_src[src_address]
        if not hasattr(src_cells[0] if src_cells else None, '__iter__'):
            src_cells = [src_cells]

        min_col, min_row, _, _ = range_boundaries(dest_address)
        for r, row in enumerate(src_cells):
            for c, cell in enumerate(row):
                dst = ws_dst.cell(row=min_row + r, column=min_col + c)
                dst.value = cell.value
                if cell.has_style:
                    dst.font = cell.font.copy()
                    dst.fill = cell.fill.copy()
                    dst.border = cell.border.copy()
                    dst.alignment = cell.alignment.copy()
                    dst.number_format = cell.number_format
        wb.save(session["path"])
        return json.dumps({"success": True, "src": src_address, "dest": dest_address})
    except Exception as e:
        return json.dumps({"success": False, "error": str(e)})


# ── formatting ────────────────────────────────────────────────────────────────

def format_range(address: str, bold: bool = None, italic: bool = None,
                 font_color: str = None, bg_color: str = None,
                 font_size: float = None, font_name: str = None,
                 h_align: str = None, v_align: str = None,
                 wrap_text: bool = None,
                 border_style: str = None, border_color: str = None,
                 sheet: str = None, workbook: str = None) -> str:
    """Apply formatting to a range of cells.

    h_align: left | center | right | general
    v_align: top | center | bottom
    border_style: thin | medium | thick | dashed | dotted | none
    """
    try:
        name, session, wb, ws = _get_context(workbook, sheet)
        OpenpyxlSnapshotUndo.push(name, session)

        border_side = _side(border_style, border_color) if border_style and border_style != "none" else None
        no_border = border_style == "none"

        for cell in _iter_range(ws, address):
            if bold is not None or italic is not None or font_color is not None \
                    or font_size is not None or font_name is not None:
                kw = {}
                if bold is not None:
                    kw["bold"] = bold
                if italic is not None:
                    kw["italic"] = italic
                if font_color is not None:
                    from openpyxl.styles.colors import Color
                    kw["color"] = parse_color(font_color)
                if font_size is not None:
                    kw["size"] = font_size
                if font_name is not None:
                    kw["name"] = font_name
                old = cell.font
                cell.font = Font(
                    bold=kw.get("bold", old.bold),
                    italic=kw.get("italic", old.italic),
                    color=kw.get("color", old.color.rgb if old.color else "FF000000"),
                    size=kw.get("size", old.size),
                    name=kw.get("name", old.name),
                )
            if bg_color is not None:
                cell.fill = PatternFill(
                    fill_type="solid",
                    fgColor=parse_color(bg_color),
                )
            if h_align is not None or v_align is not None or wrap_text is not None:
                old_al = cell.alignment
                cell.alignment = Alignment(
                    horizontal=h_align or old_al.horizontal,
                    vertical=v_align or old_al.vertical,
                    wrap_text=wrap_text if wrap_text is not None else old_al.wrap_text,
                )
            if border_side is not None:
                cell.border = Border(
                    left=border_side, right=border_side,
                    top=border_side, bottom=border_side,
                )
            elif no_border:
                cell.border = Border()

        wb.save(session["path"])
        return json.dumps({"success": True, "address": address, "formatted": True})
    except Exception as e:
        return json.dumps({"success": False, "error": str(e)})


def set_number_format(address: str, format_code: str,
                      sheet: str = None, workbook: str = None) -> str:
    """Set number format code for a range (e.g. '#,##0.00', 'yyyy-mm-dd')."""
    try:
        name, session, wb, ws = _get_context(workbook, sheet)
        OpenpyxlSnapshotUndo.push(name, session)
        for cell in _iter_range(ws, address):
            cell.number_format = format_code
        wb.save(session["path"])
        return json.dumps({"success": True, "address": address, "format": format_code})
    except Exception as e:
        return json.dumps({"success": False, "error": str(e)})


def get_number_format(address: str, sheet: str = None, workbook: str = None) -> str:
    """Get number format code for a cell."""
    try:
        _, _, _, ws = _get_context(workbook, sheet)
        cell = ws[address]
        if hasattr(cell, '__iter__'):
            cell = next(iter(next(iter(cell))))
        return json.dumps({"success": True, "address": address, "format": cell.number_format})
    except Exception as e:
        return json.dumps({"success": False, "error": str(e)})


# ── dimensions ────────────────────────────────────────────────────────────────

def set_row_height(row: int, height: float, sheet: str = None, workbook: str = None) -> str:
    """Set row height in points."""
    try:
        name, session, wb, ws = _get_context(workbook, sheet)
        ws.row_dimensions[row].height = height
        wb.save(session["path"])
        return json.dumps({"success": True, "row": row, "height": height})
    except Exception as e:
        return json.dumps({"success": False, "error": str(e)})


def set_column_width(col: str, width: float, sheet: str = None, workbook: str = None) -> str:
    """Set column width. *col* is a letter (A, B, …) or 1-based integer."""
    try:
        name, session, wb, ws = _get_context(workbook, sheet)
        col_letter = get_column_letter(int(col)) if str(col).isdigit() else col.upper()
        ws.column_dimensions[col_letter].width = width
        wb.save(session["path"])
        return json.dumps({"success": True, "col": col_letter, "width": width})
    except Exception as e:
        return json.dumps({"success": False, "error": str(e)})


def autofit(address: str = None, sheet: str = None, workbook: str = None) -> str:
    """Auto-fit column widths to content (estimate — exact requires Excel COM)."""
    try:
        name, session, wb, ws = _get_context(workbook, sheet)
        # Estimate: max character count per column × 1.2
        col_widths: dict[str, int] = {}
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is not None:
                    col_ltr = get_column_letter(cell.column)
                    length = len(str(cell.value))
                    col_widths[col_ltr] = max(col_widths.get(col_ltr, 0), length)
        for col_ltr, width in col_widths.items():
            ws.column_dimensions[col_ltr].width = min(width * 1.2 + 2, 60)
        wb.save(session["path"])
        return json.dumps({"success": True, "autofit": "estimated", "columns": list(col_widths.keys())})
    except Exception as e:
        return json.dumps({"success": False, "error": str(e)})


def hide_rows(start_row: int, end_row: int, sheet: str = None, workbook: str = None) -> str:
    """Hide rows from *start_row* to *end_row* (1-based, inclusive)."""
    try:
        name, session, wb, ws = _get_context(workbook, sheet)
        for r in range(start_row, end_row + 1):
            ws.row_dimensions[r].hidden = True
        wb.save(session["path"])
        return json.dumps({"success": True, "hidden_rows": [start_row, end_row]})
    except Exception as e:
        return json.dumps({"success": False, "error": str(e)})


def show_rows(start_row: int, end_row: int, sheet: str = None, workbook: str = None) -> str:
    """Unhide rows from *start_row* to *end_row*."""
    try:
        name, session, wb, ws = _get_context(workbook, sheet)
        for r in range(start_row, end_row + 1):
            ws.row_dimensions[r].hidden = False
        wb.save(session["path"])
        return json.dumps({"success": True, "shown_rows": [start_row, end_row]})
    except Exception as e:
        return json.dumps({"success": False, "error": str(e)})


def hide_columns(start_col: int, end_col: int, sheet: str = None, workbook: str = None) -> str:
    """Hide columns (1-based indices, inclusive)."""
    try:
        name, session, wb, ws = _get_context(workbook, sheet)
        for c in range(start_col, end_col + 1):
            ws.column_dimensions[get_column_letter(c)].hidden = True
        wb.save(session["path"])
        return json.dumps({"success": True, "hidden_cols": [start_col, end_col]})
    except Exception as e:
        return json.dumps({"success": False, "error": str(e)})


def show_columns(start_col: int, end_col: int, sheet: str = None, workbook: str = None) -> str:
    """Unhide columns (1-based indices, inclusive)."""
    try:
        name, session, wb, ws = _get_context(workbook, sheet)
        for c in range(start_col, end_col + 1):
            ws.column_dimensions[get_column_letter(c)].hidden = False
        wb.save(session["path"])
        return json.dumps({"success": True, "shown_cols": [start_col, end_col]})
    except Exception as e:
        return json.dumps({"success": False, "error": str(e)})


# ── navigation / search ───────────────────────────────────────────────────────

def freeze_panes(cell: str, sheet: str = None, workbook: str = None) -> str:
    """Freeze rows/columns above and to the left of *cell* (e.g. 'B2')."""
    try:
        name, session, wb, ws = _get_context(workbook, sheet)
        ws.freeze_panes = cell
        wb.save(session["path"])
        return json.dumps({"success": True, "freeze_panes": cell})
    except Exception as e:
        return json.dumps({"success": False, "error": str(e)})


def unfreeze_panes(sheet: str = None, workbook: str = None) -> str:
    """Remove freeze panes."""
    try:
        name, session, wb, ws = _get_context(workbook, sheet)
        ws.freeze_panes = None
        wb.save(session["path"])
        return json.dumps({"success": True, "freeze_panes": None})
    except Exception as e:
        return json.dumps({"success": False, "error": str(e)})


def find_replace(find: str, replace: str, sheet: str = None,
                 workbook: str = None, match_case: bool = False) -> str:
    """Find and replace text values across a sheet."""
    try:
        name, session, wb, ws = _get_context(workbook, sheet)
        OpenpyxlSnapshotUndo.push(name, session)
        count = 0
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is not None and isinstance(cell.value, str):
                    orig = cell.value
                    if match_case:
                        new = orig.replace(find, replace)
                    else:
                        new = re.sub(re.escape(find), replace, orig, flags=re.IGNORECASE)
                    if new != orig:
                        cell.value = new
                        count += 1
        wb.save(session["path"])
        return json.dumps({"success": True, "replaced": count, "find": find, "replace": replace})
    except Exception as e:
        return json.dumps({"success": False, "error": str(e)})


def get_used_range(sheet: str = None, workbook: str = None) -> str:
    """Return the used range address of a worksheet."""
    try:
        _, _, _, ws = _get_context(workbook, sheet)
        if ws.max_row is None or ws.max_column is None:
            return json.dumps({"success": True, "address": None, "rows": 0, "cols": 0})
        addr = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
        return json.dumps({"success": True, "address": addr,
                           "rows": ws.max_row, "cols": ws.max_column})
    except Exception as e:
        return json.dumps({"success": False, "error": str(e)})


def get_range_info(address: str, sheet: str = None, workbook: str = None) -> str:
    """Return metadata about a range (dimensions, first/last cell values)."""
    try:
        _, _, _, ws = _get_context(workbook, sheet)
        min_col, min_row, max_col, max_row = range_boundaries(address)
        first = ws.cell(min_row, min_col).value
        last = ws.cell(max_row, max_col).value
        return json.dumps({
            "success": True,
            "address": address,
            "rows": max_row - min_row + 1,
            "cols": max_col - min_col + 1,
            "first_cell": first,
            "last_cell": last,
        })
    except Exception as e:
        return json.dumps({"success": False, "error": str(e)})


def sort_range(address: str, key_col: int = 1, descending: bool = False,
               has_header: bool = True, sheet: str = None, workbook: str = None) -> str:
    """Sort a range by a key column.

    *key_col*: 1-based column offset within the range.
    *has_header*: skip the first row when sorting.
    """
    try:
        name, session, wb, ws = _get_context(workbook, sheet)
        OpenpyxlSnapshotUndo.push(name, session)
        min_col, min_row, max_col, max_row = range_boundaries(address)
        data_start = min_row + (1 if has_header else 0)
        rows = []
        for r in range(data_start, max_row + 1):
            row_data = [ws.cell(r, c).value for c in range(min_col, max_col + 1)]
            rows.append(row_data)
        key_idx = key_col - 1
        rows.sort(key=lambda row: (row[key_idx] is None, row[key_idx]), reverse=descending)
        for r_idx, row_data in enumerate(rows):
            for c_idx, val in enumerate(row_data):
                ws.cell(data_start + r_idx, min_col + c_idx, value=val)
        wb.save(session["path"])
        return json.dumps({"success": True, "sorted": address, "rows_sorted": len(rows)})
    except Exception as e:
        return json.dumps({"success": False, "error": str(e)})


# ── data validation ───────────────────────────────────────────────────────────

def add_validation(address: str, validation_type: str, formula1: str = None,
                   formula2: str = None, operator: str = "between",
                   error_message: str = None, sheet: str = None,
                   workbook: str = None) -> str:
    """Add data validation to a range.

    validation_type: list | whole | decimal | date | text_length | custom
    """
    try:
        name, session, wb, ws = _get_context(workbook, sheet)
        type_map = {
            "list": "list", "whole": "whole", "decimal": "decimal",
            "date": "date", "text_length": "textLength", "custom": "custom",
        }
        dv = DataValidation(
            type=type_map.get(validation_type, validation_type),
            formula1=formula1,
            formula2=formula2,
            operator=operator,
            showErrorMessage=bool(error_message),
            error=error_message,
        )
        dv.sqref = address
        ws.add_data_validation(dv)
        wb.save(session["path"])
        return json.dumps({"success": True, "address": address, "type": validation_type})
    except Exception as e:
        return json.dumps({"success": False, "error": str(e)})


def remove_validation(address: str = None, sheet: str = None, workbook: str = None) -> str:
    """Remove all data validations from a sheet (address filter not yet supported)."""
    try:
        name, session, wb, ws = _get_context(workbook, sheet)
        ws.data_validations.dataValidation.clear()
        wb.save(session["path"])
        return json.dumps({"success": True, "removed": "all validations"})
    except Exception as e:
        return json.dumps({"success": False, "error": str(e)})


# ── hyperlinks ────────────────────────────────────────────────────────────────

def add_hyperlink(address: str, url: str, display_text: str = None,
                  sheet: str = None, workbook: str = None) -> str:
    """Add a hyperlink to a cell."""
    try:
        name, session, wb, ws = _get_context(workbook, sheet)
        cell = ws[address]
        if hasattr(cell, '__iter__'):
            cell = next(iter(next(iter(cell))))
        cell.hyperlink = url
        if display_text is not None:
            cell.value = display_text
        cell.style = "Hyperlink"
        wb.save(session["path"])
        return json.dumps({"success": True, "address": address, "url": url})
    except Exception as e:
        return json.dumps({"success": False, "error": str(e)})


def remove_hyperlinks(address: str = None, sheet: str = None, workbook: str = None) -> str:
    """Remove hyperlinks from a range (or all hyperlinks if address is None)."""
    try:
        name, session, wb, ws = _get_context(workbook, sheet)
        if address:
            for cell in _iter_range(ws, address):
                cell.hyperlink = None
        else:
            for row in ws.iter_rows():
                for cell in row:
                    if cell.hyperlink:
                        cell.hyperlink = None
        wb.save(session["path"])
        return json.dumps({"success": True, "removed_hyperlinks": address or "all"})
    except Exception as e:
        return json.dumps({"success": False, "error": str(e)})


# ── autofilter ────────────────────────────────────────────────────────────────

def apply_range_autofilter(address: str, sheet: str = None, workbook: str = None) -> str:
    """Enable AutoFilter on a range."""
    try:
        name, session, wb, ws = _get_context(workbook, sheet)
        ws.auto_filter.ref = address
        wb.save(session["path"])
        return json.dumps({"success": True, "autofilter": address})
    except Exception as e:
        return json.dumps({"success": False, "error": str(e)})


def clear_range_autofilter(sheet: str = None, workbook: str = None) -> str:
    """Remove AutoFilter from a sheet."""
    try:
        name, session, wb, ws = _get_context(workbook, sheet)
        ws.auto_filter.ref = None
        wb.save(session["path"])
        return json.dumps({"success": True, "autofilter": None})
    except Exception as e:
        return json.dumps({"success": False, "error": str(e)})


def toggle_autofilter(sheet: str = None, workbook: str = None) -> str:
    """Toggle AutoFilter on the used range."""
    try:
        _, _, _, ws = _get_context(workbook, sheet)
        if ws.auto_filter.ref:
            return clear_range_autofilter(sheet=sheet, workbook=workbook)
        addr = f"A1:{get_column_letter(ws.max_column)}1"
        return apply_range_autofilter(addr, sheet=sheet, workbook=workbook)
    except Exception as e:
        return json.dumps({"success": False, "error": str(e)})


# ── stubs for COM-only operations ─────────────────────────────────────────────

def _com_only(op_name: str) -> str:
    return json.dumps({
        "success": False,
        "error": f"{op_name} requires Windows + Excel (COM mode).",
    })


def set_cell_lock(address: str, locked: bool = True, sheet: str = None, workbook: str = None) -> str:
    return _com_only("set_cell_lock")


def group_rows(start_row: int, end_row: int, sheet: str = None, workbook: str = None) -> str:
    return _com_only("group_rows")


def ungroup_rows(start_row: int, end_row: int, sheet: str = None, workbook: str = None) -> str:
    return _com_only("ungroup_rows")


def group_columns(start_col: int, end_col: int, sheet: str = None, workbook: str = None) -> str:
    return _com_only("group_columns")


def ungroup_columns(start_col: int, end_col: int, sheet: str = None, workbook: str = None) -> str:
    return _com_only("ungroup_columns")


def remove_duplicates(address: str, columns: list[int] = None,
                      sheet: str = None, workbook: str = None) -> str:
    return _com_only("remove_duplicates")


def get_current_region(address: str, sheet: str = None, workbook: str = None) -> str:
    return _com_only("get_current_region")
