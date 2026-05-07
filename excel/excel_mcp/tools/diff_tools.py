"""Excel diff tool — compare two workbook files (cross-platform, no Office needed)."""
import json
from pathlib import Path
from typing import Optional

import openpyxl
from openpyxl.utils import get_column_letter


def excel_diff(
    path1: str,
    path2: str,
    sheet: str = None,
    max_diffs: int = 200,
) -> str:
    """Compare two Excel workbook files and return cell-level differences.

    Works cross-platform with no Excel installation required.

    Args:
        path1: Path to the "before" workbook (or snapshot).
        path2: Path to the "after" workbook.
        sheet:  Compare only this sheet name. If None, compares all sheets
                that exist in both workbooks.
        max_diffs: Cap on returned differences (avoids huge payloads).

    Returns JSON with:
        - added_sheets: sheets in path2 but not path1
        - removed_sheets: sheets in path1 but not path2
        - changed_cells: list of {sheet, cell, before, after}
        - summary: counts
    """
    try:
        p1, p2 = Path(path1).resolve(), Path(path2).resolve()
        if not p1.exists():
            return json.dumps({"success": False, "error": f"File not found: {path1}"})
        if not p2.exists():
            return json.dumps({"success": False, "error": f"File not found: {path2}"})

        wb1 = openpyxl.load_workbook(str(p1), data_only=True, read_only=True)
        wb2 = openpyxl.load_workbook(str(p2), data_only=True, read_only=True)

        sheets1 = set(wb1.sheetnames)
        sheets2 = set(wb2.sheetnames)

        if sheet:
            compare_sheets = [sheet] if sheet in sheets1 and sheet in sheets2 else []
            missing = []
            if sheet not in sheets1:
                missing.append(f"'{sheet}' not in {path1}")
            if sheet not in sheets2:
                missing.append(f"'{sheet}' not in {path2}")
            if missing:
                return json.dumps({"success": False, "error": "; ".join(missing)})
        else:
            compare_sheets = sorted(sheets1 & sheets2)

        added_sheets = sorted(sheets2 - sheets1)
        removed_sheets = sorted(sheets1 - sheets2)
        changed_cells = []
        truncated = False

        for sh_name in compare_sheets:
            ws1 = wb1[sh_name]
            ws2 = wb2[sh_name]

            # Build value maps: (row, col) -> value
            def _map(ws):
                m = {}
                for row in ws.iter_rows():
                    for cell in row:
                        v = cell.value
                        if v is not None:
                            m[(cell.row, cell.column)] = v
                return m

            map1 = _map(ws1)
            map2 = _map(ws2)
            all_coords = set(map1) | set(map2)

            for coord in sorted(all_coords):
                v1 = map1.get(coord)
                v2 = map2.get(coord)
                if v1 != v2:
                    r, c = coord
                    cell_addr = f"{get_column_letter(c)}{r}"
                    changed_cells.append({
                        "sheet": sh_name,
                        "cell": cell_addr,
                        "before": _serialize(v1),
                        "after": _serialize(v2),
                    })
                    if len(changed_cells) >= max_diffs:
                        truncated = True
                        break
            if truncated:
                break

        wb1.close()
        wb2.close()

        return json.dumps({
            "success": True,
            "path1": str(p1),
            "path2": str(p2),
            "added_sheets": added_sheets,
            "removed_sheets": removed_sheets,
            "changed_cells": changed_cells,
            "truncated": truncated,
            "summary": {
                "added_sheets": len(added_sheets),
                "removed_sheets": len(removed_sheets),
                "changed_cells": len(changed_cells),
                "sheets_compared": len(compare_sheets),
            },
        })
    except Exception as e:
        return json.dumps({"success": False, "error": str(e)})


def _serialize(v):
    """Make a cell value JSON-safe."""
    if v is None:
        return None
    if isinstance(v, (int, float, bool, str)):
        return v
    # datetime, date, etc.
    try:
        return str(v)
    except Exception:
        return repr(v)
