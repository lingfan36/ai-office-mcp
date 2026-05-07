"""openpyxl session manager — cross-platform Excel backend (no Office required)."""
import json
import os
import shutil
import uuid
from datetime import datetime
from pathlib import Path
from typing import Optional

import openpyxl
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.utils.cell import coordinate_to_tuple, range_boundaries

SNAPSHOT_DIR = Path(os.environ.get("EXCEL_SNAPSHOT_DIR", Path.home() / ".excel_mcp" / "snapshots"))
MAX_SNAPSHOTS = int(os.environ.get("EXCEL_MAX_SNAPSHOTS", "20"))
_INDEX_FILE = SNAPSHOT_DIR / "undo_index_openpyxl.json"


# ── Session manager ────────────────────────────────────────────────────────────

class OpenpyxlSession:
    """Manages open workbooks in openpyxl mode (mirrors COM session API)."""

    # name -> {"wb": Workbook, "path": str, "read_only": bool}
    _sessions: dict[str, dict] = {}

    @classmethod
    def open(cls, path: str, read_only: bool = False) -> dict:
        abs_path = str(Path(path).resolve())
        wb = load_workbook(abs_path, read_only=read_only, data_only=False)
        name = Path(abs_path).name
        # If same name already open, close old one first
        if name in cls._sessions:
            try:
                cls._sessions[name]["wb"].close()
            except Exception:
                pass
        cls._sessions[name] = {"wb": wb, "path": abs_path, "read_only": read_only}
        return {"name": name, "path": abs_path, "sheets": len(wb.sheetnames)}

    @classmethod
    def create(cls, path: str) -> dict:
        abs_path = str(Path(path).resolve())
        wb = Workbook()
        Path(abs_path).parent.mkdir(parents=True, exist_ok=True)
        wb.save(abs_path)
        name = Path(abs_path).name
        cls._sessions[name] = {"wb": wb, "path": abs_path, "read_only": False}
        return {"name": name, "path": abs_path, "sheets": len(wb.sheetnames)}

    @classmethod
    def get(cls, name_or_path: str = None) -> tuple[str, dict]:
        """Return (name, session) or raise ValueError."""
        if not cls._sessions:
            raise ValueError("No workbook open. Call open_workbook first.")
        if name_or_path is None:
            name = next(iter(cls._sessions))
            return name, cls._sessions[name]
        target = name_or_path.lower()
        for name, session in cls._sessions.items():
            if name.lower() == target or session["path"].lower() == target:
                return name, session
        raise ValueError(f"Workbook not found: {name_or_path!r}")

    @classmethod
    def save(cls, name_or_path: str = None, save_as: str = None) -> str:
        name, session = cls.get(name_or_path)
        target = str(Path(save_as).resolve()) if save_as else session["path"]
        session["wb"].save(target)
        if save_as:
            session["path"] = target
        return target

    @classmethod
    def close(cls, name_or_path: str = None, save: bool = True) -> str:
        name, session = cls.get(name_or_path)
        if save and not session["read_only"]:
            session["wb"].save(session["path"])
        try:
            session["wb"].close()
        except Exception:
            pass
        del cls._sessions[name]
        return name

    @classmethod
    def list_all(cls) -> list[dict]:
        return [
            {
                "name": name,
                "path": s["path"],
                "read_only": s["read_only"],
                "sheets": len(s["wb"].sheetnames),
                "active": True,
            }
            for name, s in cls._sessions.items()
        ]

    @classmethod
    def get_sheet(cls, wb: Workbook, sheet_name: str = None):
        if sheet_name is None:
            return wb.active
        if sheet_name in wb.sheetnames:
            return wb[sheet_name]
        raise ValueError(f"Sheet not found: {sheet_name!r}")


# ── Snapshot-based undo (openpyxl mode) ───────────────────────────────────────

class OpenpyxlSnapshotUndo:
    """File-copy snapshot undo for openpyxl sessions."""

    _stack: list[dict] = []
    _loaded: bool = False

    @classmethod
    def _ensure_loaded(cls):
        if cls._loaded:
            return
        cls._loaded = True
        if _INDEX_FILE.exists():
            try:
                data = json.loads(_INDEX_FILE.read_text(encoding="utf-8"))
                cls._stack = [e for e in data if Path(e["snapshot"]).exists()]
            except Exception:
                cls._stack = []

    @classmethod
    def _save_index(cls):
        try:
            SNAPSHOT_DIR.mkdir(parents=True, exist_ok=True)
            _INDEX_FILE.write_text(
                json.dumps(cls._stack, indent=2, ensure_ascii=False),
                encoding="utf-8",
            )
        except Exception:
            pass

    @classmethod
    def push(cls, name: str, session: dict) -> str:
        cls._ensure_loaded()
        SNAPSHOT_DIR.mkdir(parents=True, exist_ok=True)
        snap_id = f"{datetime.now().strftime('%Y%m%d_%H%M%S')}_{uuid.uuid4().hex[:6]}"
        snap_path = str(SNAPSHOT_DIR / f"opx_{snap_id}.xlsx")
        session["wb"].save(snap_path)
        cls._stack.append({
            "id": snap_id,
            "wb_name": name,
            "wb_path": session["path"],
            "snapshot": snap_path,
        })
        while len(cls._stack) > MAX_SNAPSHOTS:
            old = cls._stack.pop(0)
            try:
                os.remove(old["snapshot"])
            except OSError:
                pass
        cls._save_index()
        return snap_id

    @classmethod
    def pop(cls, workbook_name: str = None) -> dict:
        cls._ensure_loaded()
        if not cls._stack:
            return {"success": False, "message": "Undo stack is empty"}

        idx = None
        for i in range(len(cls._stack) - 1, -1, -1):
            e = cls._stack[i]
            if workbook_name is None or e["wb_name"].lower() == workbook_name.lower():
                idx = i
                break
        if idx is None:
            return {"success": False, "message": f"No snapshot for: {workbook_name!r}"}

        entry = cls._stack.pop(idx)
        snap_path = entry["snapshot"]
        wb_path = entry["wb_path"]
        wb_name = entry["wb_name"]

        # Close current in-memory version, restore from snapshot, reopen
        if wb_name in OpenpyxlSession._sessions:
            try:
                OpenpyxlSession._sessions[wb_name]["wb"].close()
            except Exception:
                pass
        shutil.copy2(snap_path, wb_path)
        try:
            os.remove(snap_path)
        except OSError:
            pass
        wb = load_workbook(wb_path)
        OpenpyxlSession._sessions[wb_name] = {"wb": wb, "path": wb_path, "read_only": False}
        cls._save_index()
        return {"success": True, "restored": wb_path, "snapshot_id": entry["id"]}

    @classmethod
    def list(cls, workbook_name: str = None) -> list[dict]:
        cls._ensure_loaded()
        stack = cls._stack
        if workbook_name:
            stack = [s for s in stack if s["wb_name"].lower() == workbook_name.lower()]
        return [{"id": s["id"], "wb_name": s["wb_name"], "timestamp": s["id"][:15]} for s in stack]

    @classmethod
    def clear(cls, workbook_name: str = None):
        cls._ensure_loaded()
        if workbook_name:
            keep, remove = [], []
            for s in cls._stack:
                (remove if s["wb_name"].lower() == workbook_name.lower() else keep).append(s)
            cls._stack = keep
        else:
            remove = list(cls._stack)
            cls._stack = []
        for s in remove:
            try:
                os.remove(s["snapshot"])
            except OSError:
                pass
        cls._save_index()


# ── Helpers shared by compat tools ────────────────────────────────────────────

def parse_color(hex_str: str) -> str:
    """Normalise #RRGGBB → 'FFRRGGBB' (openpyxl ARGB format)."""
    h = hex_str.lstrip("#")
    if len(h) == 6:
        return "FF" + h.upper()
    return h.upper()


def range_to_rows_cols(address: str):
    """Return (min_row, min_col, max_row, max_col) from an A1-style address."""
    return range_boundaries(address)


def col_letter(n: int) -> str:
    return get_column_letter(n)


def col_index(letter: str) -> int:
    return column_index_from_string(letter)
