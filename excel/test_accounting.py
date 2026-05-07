"""
会计专业操作测试 — 覆盖资产负债表、损益表、DCF、明细账四大场景。
测试: 会计格式、财务公式(NPV/IRR/PMT/SUMIF/VLOOKUP)、条件格式、数据验证、
      命名区域、合并单元格、排序、查找替换、工作表保护、期间复制。
"""
import sys, os, json, time, tempfile
sys.path.insert(0, os.path.dirname(__file__))
os.environ.setdefault("EXCEL_SNAPSHOT_DIR", os.path.join(os.path.dirname(__file__), ".snapshots"))
os.environ.setdefault("EXCEL_MAX_SNAPSHOTS", "50")

import openpyxl
from excel_mcp.tools.file_tools      import open_workbook, save_workbook, close_workbook
from excel_mcp.tools.range_tools     import (
    set_values, set_formulas, format_range, set_number_format,
    merge_cells, autofit, add_validation, find_replace,
    sort_range, get_values, set_cell_lock,
)
from excel_mcp.tools.sheet_tools     import (
    create_sheet, rename_sheet, copy_sheet, set_tab_color,
    protect_sheet, unprotect_sheet,
)
from excel_mcp.tools.named_range_tools import create_named_range, read_named_range, list_named_ranges
from excel_mcp.tools.format_tools    import add_conditional_format
from excel_mcp.tools.table_tools     import create_table, get_table_data, apply_table_style
from excel_mcp.tools.pivot_tools     import create_pivot_table, add_pivot_field, refresh_pivot, list_pivot_fields
from excel_mcp.tools.chart_tools     import create_chart
from excel_mcp.tools.calc_tools      import calculate
from excel_mcp.tools.undo_tools      import list_undo_snapshots

results = []

def check(label, raw):
    d = json.loads(raw) if isinstance(raw, str) else raw
    ok = d.get("success", False)
    print(f"  {'OK' if ok else 'NG'}  {label}")
    if not ok:
        print(f"       ERR: {d.get('error', d)}")
    results.append((label, ok))
    return d

# ── 准备空白工作簿 ─────────────────────────────────────────────────────────────
print("\n[ 0 ] 准备工作簿")

# Kill any stale WPS instances so we start with a clean COM state.
# This avoids NUIDialog blocking Workbooks.Count in the cleanup loop below.
import subprocess as _sp
_sp.run(["taskkill", "/f", "/im", "et.exe"],  capture_output=True)
_sp.run(["taskkill", "/f", "/im", "wps.exe"], capture_output=True)
time.sleep(4)

xlsx_path = os.path.join(tempfile.gettempdir(), f"acct_test_{int(time.time())}.xlsx")
wb0 = openpyxl.Workbook(); wb0.active.title = "Sheet1"; wb0.save(xlsx_path)
check("open_workbook", open_workbook(xlsx_path))
time.sleep(1)

# 建立所需工作表
check("rename Sheet1 -> 资产负债表", rename_sheet("Sheet1", "资产负债表"))
check("create 损益表",  create_sheet("损益表"))
check("create DCF模型", create_sheet("DCF模型"))
check("create 明细账",  create_sheet("明细账"))
check("create 辅助表",  create_sheet("辅助表"))

# ══════════════════════════════════════════════════════════════════════════════
print("\n[ 1 ] 资产负债表 —— 格式、公式、命名区域、合并、保护")
# ── 标题 ──────────────────────────────────────────────────────────────────────
check("merge title",
    merge_cells("A1:F1", sheet="资产负债表"))
check("set title text",
    set_values("A1", [["XX 科技有限公司  资产负债表  2025-12-31"]], sheet="资产负债表"))
check("format title",
    format_range("A1:F1", sheet="资产负债表",
                 bold=True, font_size=14, bg_color="#1F4E79", font_color="#FFFFFF",
                 h_align="center", v_align="center"))

# ── 表头行 ────────────────────────────────────────────────────────────────────
headers = [["资产", "期末余额", "期初余额", "负债及所有者权益", "期末余额", "期初余额"]]
check("set col headers", set_values("A2", headers, sheet="资产负债表"))
check("format col headers",
    format_range("A2:F2", sheet="资产负债表",
                 bold=True, bg_color="#2E75B6", font_color="#FFFFFF",
                 h_align="center", border="outer"))

# ── 资产数据 (B列=期末, C列=期初) ─────────────────────────────────────────────
assets = [
    ["流动资产：", None, None, "流动负债：", None, None],
    ["  货币资金",        850000,  720000, "  短期借款",         200000,  150000],
    ["  应收账款",        430000,  380000, "  应付账款",         310000,  270000],
    ["  存货",            280000,  260000, "  预收款项",          45000,   38000],
    ["  预付账款",         35000,   28000, "  应付职工薪酬",       82000,   76000],
    ["  其他流动资产",     15000,   12000, "  应交税费",          28000,   24000],
    ["流动资产合计",     None,    None,   "流动负债合计",       None,    None],  # row9 = A9
    [None, None, None, None, None, None],
    ["非流动资产：", None, None, "非流动负债：", None, None],
    ["  固定资产净值",    1200000, 1350000, "  长期借款",         500000,  600000],
    ["  无形资产",         180000,  200000, "  长期应付款",        80000,   90000],
    ["  长期投资",         250000,  220000, None, None, None],
    ["非流动资产合计",   None,    None,   "非流动负债合计",     None,    None],  # row14
    [None, None, None, None, None, None],
    [None, None, None, "所有者权益：", None, None],
    [None, None, None, "  实收资本",           800000,  800000],
    [None, None, None, "  资本公积",           200000,  200000],
    [None, None, None, "  未分配利润",         270000,  142000],
    [None, None, None, "所有者权益合计",      None,    None],  # row20
    ["资产总计", None, None, "负债和所有者权益合计", None, None],  # row21
]
check("set assets data", set_values("A3", assets, sheet="资产负债表"))

# ── 汇总公式 (行号从3开始, 所以第7行数据是行9) ─────────────────────────────────
# 流动资产合计 B9 = B4:B8
check("formula 流动资产合计(期末)",   set_formulas("B9",  [["=SUM(B4:B8)"]], sheet="资产负债表"))
check("formula 流动资产合计(期初)",   set_formulas("C9",  [["=SUM(C4:C8)"]], sheet="资产负债表"))
# 非流动资产合计 B14
check("formula 非流动资产合计(期末)", set_formulas("B14", [["=SUM(B11:B13)"]], sheet="资产负债表"))
check("formula 非流动资产合计(期初)", set_formulas("C14", [["=SUM(C11:C13)"]], sheet="资产负债表"))
# 资产总计 B21
check("formula 资产总计(期末)",       set_formulas("B21", [["=B9+B14"]], sheet="资产负债表"))
check("formula 资产总计(期初)",       set_formulas("C21", [["=C9+C14"]], sheet="资产负债表"))
# 负债汇总
check("formula 流动负债合计(期末)",   set_formulas("E9",  [["=SUM(E4:E8)"]], sheet="资产负债表"))
check("formula 流动负债合计(期初)",   set_formulas("F9",  [["=SUM(F4:F8)"]], sheet="资产负债表"))
check("formula 非流动负债合计(期末)", set_formulas("E14", [["=SUM(E11:E12)"]], sheet="资产负债表"))
check("formula 非流动负债合计(期初)", set_formulas("F14", [["=SUM(F11:F12)"]], sheet="资产负债表"))
# 所有者权益合计
check("formula 权益合计(期末)",       set_formulas("E20", [["=SUM(E17:E19)"]], sheet="资产负债表"))
check("formula 权益合计(期初)",       set_formulas("F20", [["=SUM(F17:F19)"]], sheet="资产负债表"))
# 负债和所有者权益合计
check("formula 负债权益总计(期末)",   set_formulas("E21", [["=E9+E14+E20"]], sheet="资产负债表"))
check("formula 负债权益总计(期初)",   set_formulas("F21", [["=F9+F14+F20"]], sheet="资产负债表"))

# ── 会计数字格式 ──────────────────────────────────────────────────────────────
# 中国会计常用：负数加括号
acct_fmt = '_-* #,##0_-;-* #,##0_-;_-* "-"_-;_-@_-'
check("accounting format B:C",  set_number_format("B3:C21", acct_fmt, sheet="资产负债表"))
check("accounting format E:F",  set_number_format("E3:F21", acct_fmt, sheet="资产负债表"))

# ── 小计行加粗底部边框 ────────────────────────────────────────────────────────
for subtotal_row in ["A9:F9", "A14:F14", "A21:F21"]:
    check(f"bold subtotal {subtotal_row}",
          format_range(subtotal_row, sheet="资产负债表", bold=True, border="outer"))

# ── 命名区域 ──────────────────────────────────────────────────────────────────
check("named range TotalAssets",
      create_named_range("TotalAssets", "B21", sheet="资产负债表", comment="资产总计"))
check("named range TotalLiabEquity",
      create_named_range("TotalLiabEquity", "E21", sheet="资产负债表", comment="负债权益总计"))
r = check("list named ranges", list_named_ranges())
print(f"       named ranges count = {r.get('count')}")
r = check("read TotalAssets", read_named_range("TotalAssets"))
print(f"       TotalAssets = {r.get('value')}")

# ── 会计恒等式验证公式 (F1 区域放校验) ────────────────────────────────────────
balance_f = '=IF(ABS(TotalAssets-TotalLiabEquity)<1,"OK - 借贷平衡","ERR - 借贷不平衡")'
check("balance check formula",
      set_formulas("A23", [[balance_f]],
                   sheet="资产负债表"))
check("format balance check",
      format_range("A23:C23", sheet="资产负债表", italic=True, font_color="#375623"))
check("autofit cols",   autofit("A1:F21", target="columns", sheet="资产负债表"))

# ── 单元格锁定 + 工作表保护 ───────────────────────────────────────────────────
# 先解锁所有单元格，再锁定公式行，防止误操作
check("unlock all cells",  set_cell_lock("A1:F30", locked=False, sheet="资产负债表"))
check("lock formula cells B9",  set_cell_lock("B9:F9",   locked=True, sheet="资产负债表"))
check("lock formula cells B14", set_cell_lock("B14:F14", locked=True, sheet="资产负债表"))
check("lock formula cells B21", set_cell_lock("B21:F21", locked=True, sheet="资产负债表"))
check("lock formula cells E20", set_cell_lock("E20:F20", locked=True, sheet="资产负债表"))
check("protect sheet", protect_sheet("资产负债表", password="Acc@2025"))
check("unprotect sheet", unprotect_sheet("资产负债表", password="Acc@2025"))

# ══════════════════════════════════════════════════════════════════════════════
print("\n[ 2 ] 损益表 —— 差异分析、条件格式、百分比")
pl_headers = [["项目", "本期实际", "本期预算", "差异额", "差异率%", "上期实际"]]
pl_data = [
    ["一、营业收入",       5200000, 5000000, None, None, 4800000],
    ["  减：营业成本",     3120000, 3000000, None, None, 2880000],
    ["二、毛利润",              None, None, None, None, None],
    ["  毛利率%",               None, None, None, None, None],
    ["  减：销售费用",      520000,  480000, None, None, 450000],
    ["  减：管理费用",      380000,  360000, None, None, 340000],
    ["  减：研发费用",      210000,  200000, None, None, 180000],
    ["三、营业利润(EBIT)", None,    None,   None, None, None],
    ["  减：利息费用",       45000,   42000, None, None, 40000],
    ["四、税前利润(EBT)",  None,    None,   None, None, None],
    ["  减：所得税(25%)",  None,    None,   None, None, None],
    ["五、净利润",          None,    None,   None, None, None],
    ["  净利率%",           None,    None,   None, None, None],
]
check("set P&L headers", set_values("A1", pl_headers, sheet="损益表"))
check("set P&L data",    set_values("A2", pl_data,    sheet="损益表"))
check("format P&L header",
    format_range("A1:F1", sheet="损益表", bold=True, bg_color="#375623", font_color="#FFFFFF", h_align="center"))

# 计算公式
check("毛利润",    set_formulas("B4",  [["=B2-B3"]], sheet="损益表"))
check("毛利润预算", set_formulas("C4",  [["=C2-C3"]], sheet="损益表"))
check("毛利润上期", set_formulas("F4",  [["=F2-F3"]], sheet="损益表"))
check("毛利率%",   set_formulas("B5",  [["=IF(B2=0,0,B4/B2)"]], sheet="损益表"))
check("毛利率%预算",set_formulas("C5",  [["=IF(C2=0,0,C4/C2)"]], sheet="损益表"))
check("毛利率%上期",set_formulas("F5",  [["=IF(F2=0,0,F4/F2)"]], sheet="损益表"))
check("EBIT",      set_formulas("B9",  [["=B4-B6-B7-B8"]], sheet="损益表"))
check("EBIT预算",  set_formulas("C9",  [["=C4-C6-C7-C8"]], sheet="损益表"))
check("EBIT上期",  set_formulas("F9",  [["=F4-F6-F7-F8"]], sheet="损益表"))
check("EBT",       set_formulas("B11", [["=B9-B10"]], sheet="损益表"))
check("EBT预算",   set_formulas("C11", [["=C9-C10"]], sheet="损益表"))
check("EBT上期",   set_formulas("F11", [["=F9-F10"]], sheet="损益表"))
check("所得税",    set_formulas("B12", [["=B11*0.25"]], sheet="损益表"))
check("所得税预算",set_formulas("C12", [["=C11*0.25"]], sheet="损益表"))
check("所得税上期",set_formulas("F12", [["=F11*0.25"]], sheet="损益表"))
check("净利润",    set_formulas("B13", [["=B11-B12"]], sheet="损益表"))
check("净利润预算",set_formulas("C13", [["=C11-C12"]], sheet="损益表"))
check("净利润上期",set_formulas("F13", [["=F11-F12"]], sheet="损益表"))
check("净利率%",   set_formulas("B14", [["=IF(B2=0,0,B13/B2)"]], sheet="损益表"))
check("净利率%预算",set_formulas("C14", [["=IF(C2=0,0,C13/C2)"]], sheet="损益表"))

# 差异列 D = 实际 - 预算，E = D/C
for r in range(2, 15):
    check(f"差异额 D{r}", set_formulas(f"D{r}", [[f"=IF(AND(B{r}<>\"\",C{r}<>\"\"),B{r}-C{r},\"\")"]], sheet="损益表"))
    check(f"差异率 E{r}", set_formulas(f"E{r}", [[f"=IF(AND(D{r}<>\"\",C{r}<>0),D{r}/C{r},\"\")"]], sheet="损益表"))

# 数字格式
check("acct format B:D", set_number_format("B2:D14", acct_fmt, sheet="损益表"))
check("pct format E:E",  set_number_format("E2:E14", "0.0%", sheet="损益表"))
check("pct format B5",   set_number_format("B5:F5",  "0.0%", sheet="损益表"))
check("pct format B14",  set_number_format("B14:F14","0.0%", sheet="损益表"))

# 条件格式: 差异额 D列 正数=绿, 负数=红
check("CF 差异额 positive",
      add_conditional_format("D2:D14", "cell_value", operator="greater", value=0,
                              font_color="#375623", bg_color="#E2EFDA", sheet="损益表"))
check("CF 差异额 negative",
      add_conditional_format("D2:D14", "cell_value", operator="less", value=0,
                              font_color="#9C0006", bg_color="#FFC7CE", sheet="损益表"))
# 差异率色阶
check("CF 差异率 color_scale",
      add_conditional_format("E2:E14", "color_scale", sheet="损益表"))
# 数据条 实际收入
check("CF 实际 data_bar",
      add_conditional_format("B2:B14", "data_bar", sheet="损益表"))

check("bold subtotals P&L",
      format_range("A4:F4", sheet="损益表", bold=True, border="outer"))
for row in ["A9:F9", "A11:F11", "A13:F13"]:
    check(f"bold {row}", format_range(row, sheet="损益表", bold=True, border="outer"))
check("autofit P&L", autofit("A1:F14", target="columns", sheet="损益表"))
check("tab color 损益表", set_tab_color("损益表", "#375623"))

# ══════════════════════════════════════════════════════════════════════════════
print("\n[ 3 ] DCF模型 —— NPV / IRR / PMT / XNPV 金融函数")
# 贷款摊还表 (PMT)
loan_title = [["贷款摊还计划表"]]
loan_params = [
    ["贷款金额",   1000000],
    ["年利率",        0.045],
    ["期数(月)",        60],
    ["月还款额",      None],  # PMT公式
]
check("set loan title", set_values("A1", loan_title, sheet="DCF模型"))
check("merge loan title", merge_cells("A1:F1", sheet="DCF模型"))
check("format loan title",
    format_range("A1:F1", sheet="DCF模型", bold=True, font_size=13,
                 bg_color="#1F4E79", font_color="#FFFFFF", h_align="center"))
check("set loan params", set_values("A3", loan_params, sheet="DCF模型"))
# PMT(rate, nper, pv)
check("PMT formula", set_formulas("B6", [["=PMT(B4/12,B5,-B3)"]], sheet="DCF模型"))
check("format PMT cell", set_number_format("B3:B6", "#,##0.00", sheet="DCF模型"))
check("bold PMT label",  format_range("A6:B6", sheet="DCF模型", bold=True, bg_color="#FFF2CC"))

# 摊还明细表
check("set amort header",
    set_values("A8", [["期数","期初余额","月还款","利息","本金","期末余额"]], sheet="DCF模型"))
check("format amort header",
    format_range("A8:F8", sheet="DCF模型", bold=True, bg_color="#2E75B6",
                 font_color="#FFFFFF", h_align="center"))
# 前12期摊还
amort_formulas = []
for i in range(1, 13):
    row = 8 + i
    if i == 1:
        amort_formulas.append([
            str(i),
            f"=$B$3",                       # 期初余额
            f"=$B$6",                       # 月还款
            f"=B{row}*$B$4/12",             # 利息
            f"=C{row}-D{row}",              # 本金
            f"=B{row}-E{row}",              # 期末余额
        ])
    else:
        prev = 8 + i - 1
        amort_formulas.append([
            str(i),
            f"=F{prev}",
            f"=$B$6",
            f"=B{row}*$B$4/12",
            f"=C{row}-D{row}",
            f"=B{row}-E{row}",
        ])
check("set amort formulas", set_formulas("A9", amort_formulas, sheet="DCF模型"))
check("format amort numbers", set_number_format("B9:F20", "#,##0.00", sheet="DCF模型"))
check("format amort rows",
    format_range("A9:F20", sheet="DCF模型", border="all"))

# DCF 净现值/IRR
dcf_title = [["项目DCF估值模型 (自由现金流)"]]
check("set DCF title", set_values("H1", dcf_title, sheet="DCF模型"))
check("merge DCF title", merge_cells("H1:N1", sheet="DCF模型"))
check("format DCF title",
    format_range("H1:N1", sheet="DCF模型", bold=True, font_size=12,
                 bg_color="#833C00", font_color="#FFFFFF", h_align="center"))

dcf_params = [
    ["折现率(WACC)", 0.10],
    ["期初投资",     -500000],
]
check("set DCF params", set_values("H3", dcf_params, sheet="DCF模型"))
check("format DCF pct", set_number_format("I3", "0.0%", sheet="DCF模型"))

dcf_cf_header = [["年份", "自由现金流"]]
dcf_cf = [
    [1,  80000],
    [2, 110000],
    [3, 150000],
    [4, 200000],
    [5, 280000],
]
check("set DCF CF header", set_values("H5", dcf_cf_header, sheet="DCF模型"))
check("set DCF CF data",   set_values("H6", dcf_cf, sheet="DCF模型"))
check("format DCF CF header",
    format_range("H5:I5", sheet="DCF模型", bold=True, bg_color="#2E75B6",
                 font_color="#FFFFFF", h_align="center"))
check("format DCF CF numbers", set_number_format("I6:I10", "#,##0", sheet="DCF模型"))

check("NPV formula",    set_formulas("I12", [["=NPV(I3,I6:I10)+I4"]], sheet="DCF模型"))
check("IRR formula",    set_formulas("I13", [["=IRR(I4:I10)"]], sheet="DCF模型"))  # 含期初投资
check("set NPV label",  set_values("H12", [["NPV"]], sheet="DCF模型"))
check("set IRR label",  set_values("H13", [["IRR"]], sheet="DCF模型"))
check("format NPV",     set_number_format("I12", "#,##0", sheet="DCF模型"))
check("format IRR",     set_number_format("I13", "0.00%", sheet="DCF模型"))
check("bold NPV/IRR",   format_range("H12:I13", sheet="DCF模型", bold=True, bg_color="#FFF2CC", border="outer"))
check("autofit DCF",    autofit("A1:N20", target="columns", sheet="DCF模型"))
check("tab color DCF",  set_tab_color("DCF模型", "#833C00"))

# ══════════════════════════════════════════════════════════════════════════════
print("\n[ 4 ] 明细账 —— 数据验证、SUMIF、VLOOKUP、排序、查找替换")
gl_headers = [["凭证日期", "科目代码", "科目名称", "科目类型", "借方金额", "贷方金额", "说明"]]
gl_data = [
    ["2025-01-05", "1001", "库存现金",   "资产", 50000,  0,      "收到股东投资"],
    ["2025-01-05", "4001", "实收资本",   "权益", 0,      50000,  "收到股东投资"],
    ["2025-01-10", "1002", "银行存款",   "资产", 200000, 0,      "银行贷款到账"],
    ["2025-01-10", "2001", "短期借款",   "负债", 0,      200000, "银行贷款到账"],
    ["2025-01-15", "5001", "主营业务收入","收入", 0,      85000,  "销售产品A"],
    ["2025-01-15", "1122", "应收账款",   "资产", 85000,  0,      "销售产品A"],
    ["2025-01-20", "5401", "销售费用",   "费用", 12000,  0,      "支付销售佣金"],
    ["2025-01-20", "1002", "银行存款",   "资产", 0,      12000,  "支付销售佣金"],
    ["2025-02-03", "1401", "原材料",     "资产", 36000,  0,      "采购原材料"],
    ["2025-02-03", "2202", "应付账款",   "负债", 0,      36000,  "采购原材料"],
    ["2025-02-10", "5101", "管理费用",   "费用", 8500,   0,      "支付办公费"],
    ["2025-02-10", "1002", "银行存款",   "资产", 0,      8500,   "支付办公费"],
    ["2025-02-18", "5001", "主营业务收入","收入", 0,      120000, "销售产品B"],
    ["2025-02-18", "1122", "应收账款",   "资产", 120000, 0,      "销售产品B"],
    ["2025-03-01", "1002", "银行存款",   "资产", 85000,  0,      "收回应收账款"],
    ["2025-03-01", "1122", "应收账款",   "资产", 0,      85000,  "收回应收账款"],
]
check("set GL headers", set_values("A1", gl_headers, sheet="明细账"))
check("set GL data",    set_values("A2", gl_data, sheet="明细账"))
check("format GL header",
    format_range("A1:G1", sheet="明细账", bold=True, bg_color="#2E75B6",
                 font_color="#FFFFFF", h_align="center"))
check("format GL amounts", set_number_format("E2:F17", "#,##0.00", sheet="明细账"))
check("GL border all",     format_range("A1:G17", sheet="明细账", border="all"))

# 数据验证: 科目类型下拉列表
check("validation dropdown 科目类型",
      add_validation("D2:D17", "list", "资产,负债,权益,收入,费用",
                     prompt="请选择科目类型", sheet="明细账"))

# 整数验证: 借贷金额 >= 0
check("validation amount >=0",
      add_validation("E2:F17", "decimal", "0",
                     operator="greater_equal",
                     error_message="金额不能为负数", sheet="明细账"))

# 按日期升序排序
check("sort by date", sort_range("A1:G17", key_col=1, order="asc", has_header=True, sheet="明细账"))

# 汇总区: SUMIF 按科目类型汇总借贷方
check("set summary title",  set_values("I1", [["科目类型汇总"]], sheet="明细账"))
check("set summary headers",set_values("I2", [["科目类型","借方合计","贷方合计","净额"]], sheet="明细账"))
check("merge summary title",merge_cells("I1:L1", sheet="明细账"))
check("format summary title",
    format_range("I1:L1", sheet="明细账", bold=True, bg_color="#1F4E79", font_color="#FFFFFF", h_align="center"))
check("format summary header",
    format_range("I2:L2", sheet="明细账", bold=True, bg_color="#2E75B6", font_color="#FFFFFF", h_align="center"))

categories = ["资产", "负债", "权益", "收入", "费用"]
for i, cat in enumerate(categories):
    r = 3 + i
    check(f"summary {cat}",
          set_formulas(f"I{r}", [[cat, f'=SUMIF($D$2:$D$17,I{r},$E$2:$E$17)',
                                  f'=SUMIF($D$2:$D$17,I{r},$F$2:$F$17)',
                                  f'=J{r}-K{r}']],
                       sheet="明细账"))

# 总计行
check("summary total row",
      set_formulas("I8", [["合计",
                            "=SUM(J3:J7)", "=SUM(K3:K7)", "=SUM(L3:L7)"]],
                   sheet="明细账"))
check("format summary numbers",
      set_number_format("J3:L8", "#,##0.00", sheet="明细账"))
check("bold summary total", format_range("I8:L8", sheet="明细账", bold=True, border="outer"))

# 辅助表: 科目代码 -> 科目名称对照 (供 VLOOKUP)
check("set aux headers", set_values("A1", [["科目代码","科目名称","科目类型"]], sheet="辅助表"))
acct_map = [
    ["1001","库存现金","资产"],["1002","银行存款","资产"],
    ["1122","应收账款","资产"],["1401","原材料","资产"],
    ["2001","短期借款","负债"],["2202","应付账款","负债"],
    ["4001","实收资本","权益"],
    ["5001","主营业务收入","收入"],
    ["5101","管理费用","费用"],["5401","销售费用","费用"],
]
check("set aux data", set_values("A2", acct_map, sheet="辅助表"))
check("format aux header",
    format_range("A1:C1", sheet="辅助表", bold=True, bg_color="#2E75B6", font_color="#FFFFFF"))

# 在明细账 H 列用 VLOOKUP 验证科目名称 (与C列比对)
check("set vlookup header", set_values("H1", [["VLOOKUP校验"]], sheet="明细账"))
for i in range(2, 18):
    check(f"vlookup row {i}",
          set_formulas(f"H{i}",
              [[f'=IF(VLOOKUP(B{i},辅助表!$A:$C,2,0)=C{i},"OK","科目名不符")']],
              sheet="明细账"))
check("format vlookup col", set_number_format("H2:H17", "@", sheet="明细账"))

# 查找替换: 将旧科目代码 "1001" 批量更新为 "1000"（模拟科目重编）
check("find_replace acct code",
      find_replace("1001", "1000", sheet="明细账"))
# 再改回来
check("find_replace revert",
      find_replace("1000", "1001", sheet="明细账"))

check("autofit GL", autofit("A1:L17", target="columns", sheet="明细账"))
check("tab color 明细账", set_tab_color("明细账", "#7030A0"))

# ══════════════════════════════════════════════════════════════════════════════
print("\n[ 5 ] 跨期对比 —— 复制工作表、pivot、图表")
# 复制资产负债表用于期初对比
check("copy sheet 资产负债表 -> 上期比较",
      copy_sheet("资产负债表", "上期比较"))
check("tab color 上期比较", set_tab_color("上期比较", "#808080"))

# 明细账透视: 按科目类型和月份汇总
check("create pivot 明细账",
      create_pivot_table("A1:G17", "B3", "GLPivot",
                         source_sheet="明细账", dest_sheet="辅助表"))
time.sleep(1)
r_fields = json.loads(list_pivot_fields("GLPivot", sheet="辅助表"))
print(f"       pivot fields: {[f['name'] for f in r_fields.get('fields', [])]}")
check("pivot row 科目类型", add_pivot_field("GLPivot", "科目类型", "row", sheet="辅助表"))
check("pivot val 借方金额", add_pivot_field("GLPivot", "借方金额", "value", sheet="辅助表"))
check("pivot val 贷方金额", add_pivot_field("GLPivot", "贷方金额", "value", sheet="辅助表"))
check("refresh GLPivot",   refresh_pivot("GLPivot", "辅助表"))

# 损益表图表
check("create P&L chart",
      create_chart("B2:C14", chart_type="bar_clustered",
                   title="实际 vs 预算对比", sheet="损益表"))

# ══════════════════════════════════════════════════════════════════════════════
print("\n[ 6 ] 计算 & 快照")
check("force calculate", calculate())
r = check("list snapshots", list_undo_snapshots())
print(f"       snapshots = {r.get('count')}")

# ══════════════════════════════════════════════════════════════════════════════
print("\n[ 7 ] 保存 & 关闭")
check("tab color 资产负债表", set_tab_color("资产负债表", "#1F4E79"))
check("save_workbook",  save_workbook())
check("close_workbook", close_workbook())

# ── 汇总 ──────────────────────────────────────────────────────────────────────
passed = sum(1 for _, ok in results if ok)
total  = len(results)
print(f"\n{'='*60}")
print(f"  RESULT: {passed}/{total} passed")
if passed < total:
    print("  FAILED:")
    for label, ok in results:
        if not ok:
            print(f"    - {label}")
print(f"{'='*60}")
print(f"  Output: {xlsx_path}")
